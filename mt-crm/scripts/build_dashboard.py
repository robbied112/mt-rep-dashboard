#!/usr/bin/env python3
"""
Missing Thorn BI Dashboard - Auto-Rebuild Script
Reads 7 Excel exports from VIP/QuickBooks, extracts all data,
and regenerates the HTML dashboard with fresh numbers.

Usage:
  python3 build_dashboard.py --data ~/OneDrive/MT\ Dashboard\ Data/ --template templates/executive_dashboard.html --output dist/index.html
  python3 build_dashboard.py --base-dir /path/to/Raw\ Exports  (legacy flag, same as --data)
"""
import openpyxl
import json
import re
import sys
import os
from datetime import datetime
from collections import defaultdict

# ============================================================
# CONFIGURATION
# ============================================================
DEFAULT_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '')
DEFAULT_TEMPLATE = None  # Will look for most recent *_MT_BI_Dashboard.html
DEFAULT_OUTPUT = None     # Will create dated file + index.html

# Wine name mapping (VIP long names -> dashboard short names)
VIP_WINE_MAP = {
    'Missing Thorn Alcohol Removed Red Wine 12/750 ml': 'Still Red',
    'Missing Thorn Alcohol Removed Still White Wine 12/750 ml': 'Still White',
    'Missing Thorn Alcohol Removed Still Rose 12/750 ml': 'Still Rose',
    'Missing Thorn Alcohol Removed Sparkling White Wine 12/750 ml': 'Sparkling White',
    'Missing Thorn Alcohol Removed Sparkling Rose Wine 12/750 ml': 'Sparkling Rose',
}
WINE_ORDER = ['Still Red', 'Still White', 'Still Rose', 'Sparkling White', 'Sparkling Rose']

# Distributors whose pricing is NOT reported (show dashes instead of misleading rev/rpc)
# Use partial matching - if any of these substrings appear in the distributor name, flag as noPrice
NO_PRICE_PATTERNS = [
    'Republic',  # Republic National Dist - all TX locations have unreliable pricing
]
# More specific: only flag these specific locations
NO_PRICE_DISTRIBUTORS = [
    'Dallas',
    'Houston',
]

# Short names for placement data
DIST_SHORT_NAMES = {
    'Breakthru Beverage Florida - Miramar': 'Breakthru FL Mir',
    'Breakthru Beverage Florida - Riverview': 'Breakthru FL Riv',
    'Breakthru Beverage South Carolina - Columbia': 'Breakthru SC',
    'Breakthru Beverage Connecticut - North Haven': 'Breakthru CT',
    'Breakthru Beverage Nevada - Las Vegas': 'Breakthru NV LV',
    'Breakthru Beverage Nevada - Reno': 'Breakthru NV Reno',
    'Fedway Associates - Kearny': 'Fedway NJ',
    'Republic National Distributing Company - Dallas/FTW': 'Republic DAL TX',
    'Republic National Distributing Company - Houston': 'Republic HOU TX',
    'Republic National Distributing Company - San Antonio': 'Republic SA TX',
    'Republic National Distributing Company - Atlanta': 'Republic ATL',
    'Tryon Distributing Company - Durham': 'Tryon Durham NC',
    'Tryon Distributing Co': 'Tryon Co NC',
    'Quail Distributing Company': 'Quail AZ',
    'Pine State Trading Co': 'Pine State ME',
    'Shangy Inc.': 'Shangy PA',
}

# Multi-warehouse distributor groups -- these share inventory within a state
# Key = parent display name, members = prefixes to match cleaned dist names
MULTI_WAREHOUSE_GROUPS = {
    'Breakthru FL (Combined)': {
        'st': 'FL',
        'members': ['Breakthru FL - Miramar', 'Breakthru FL - Riverview'],
    },
    'RNDC TX (Combined)': {
        'st': 'TX',
        'members': ['Republic National Dist - Dallas', 'Republic National Dist - Houston',
                     'Republic National Dist - San Antonio'],
    },
    'Winebow CA (Combined)': {
        'st': 'CA',
        'members': ['Winebow - CA North', 'Winebow - CA South'],
    },
}

# Reverse lookup: build at import time
_DIST_TO_GROUP = {}
for _grp_name, _grp in MULTI_WAREHOUSE_GROUPS.items():
    for _member in _grp['members']:
        _DIST_TO_GROUP[_member] = _grp_name

def get_warehouse_group(dist_name):
    """Return parent group name if dist is in a multi-warehouse group, else None."""
    for prefix, grp_name in _DIST_TO_GROUP.items():
        if dist_name.startswith(prefix):
            return grp_name
    return None

# State name lookup
STATE_NAMES = {
    'NJ': 'New Jersey', 'FL': 'Florida', 'PA': 'Pennsylvania',
    'TX': 'Texas', 'NC': 'North Carolina', 'AZ': 'Arizona',
    'GA': 'Georgia', 'NV': 'Nevada', 'CT': 'Connecticut',
    'ME': 'Maine', 'SC': 'South Carolina', 'CA': 'California',
}

# Region mapping for East/West filter
REGION_MAP = {
    'NJ': 'East', 'FL': 'East', 'PA': 'East', 'NC': 'East',
    'GA': 'East', 'CT': 'East', 'ME': 'East', 'SC': 'East',
    'TX': 'West', 'AZ': 'West', 'NV': 'West', 'CA': 'West',
}

# ============================================================
# UTILITY FUNCTIONS
# ============================================================
def clean_dist(name):
    if not name or name == 'Total':
        return name
    if 'Report Created' in str(name):
        return None
    return str(name).strip()

def map_wine(vip_name):
    return VIP_WINE_MAP.get(str(vip_name), str(vip_name))

def safe_num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def safe_pct(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return None

def fmt_yoy(prior, current):
    if prior is None or prior < 2.0:
        # Treat prior < 2.0 CE as effectively "New" -- tiny prior values
        # (e.g. 0.3 CE) produce misleading 1000%+ growth percentages
        return 'New'
    pct = ((current - prior) / abs(prior)) * 100
    if pct >= 0:
        return f'+{pct:.1f}%'
    return f'{pct:.1f}%'

def fmt_mom(prior_w4, current_w4):
    if prior_w4 is None or prior_w4 < 1.0:
        # Treat prior < 1.0 CE as effectively "New" for 4W comparison
        return 'New'
    pct = ((current_w4 - prior_w4) / abs(prior_w4)) * 100
    if pct >= 0:
        return f'+{pct:.1f}%'
    return f'{pct:.1f}%'

def js_str(s):
    """Escape a string for JS single-quoted literal."""
    return str(s).replace("\\", "\\\\").replace("'", "\\'")

def parse_date_str(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), '%m/%d/%Y')
    except:
        return None

def get_dist_short(dist_name):
    """Get short distributor name for placement charts."""
    for long, short in DIST_SHORT_NAMES.items():
        if long.lower() in dist_name.lower() or dist_name.lower() in long.lower():
            return short
    # Fallback: first 20 chars
    return dist_name[:20] if len(dist_name) > 20 else dist_name

def get_state_from_dist(dist_name):
    """Extract state abbreviation from distributor name suffix."""
    parts = dist_name.split(',')
    if len(parts) >= 2:
        st = parts[-1].strip().upper()
        if len(st) == 2:
            return st
    # Check for known state patterns
    for st, name in STATE_NAMES.items():
        if f', {st}' in dist_name or f'- {name}' in dist_name:
            return st
    return ''

def is_sample_account(name):
    """Detect sample/marketing accounts in VIP data.
    Rule: account name contains 'SAMPLE' (case-insensitive)."""
    return 'sample' in str(name).lower()


def is_likely_personal_sample(name, revenue):
    """Detect likely personal/sample entries in QB data.
    Rule: if name looks like a person (2-3 words, all alpha, no business keywords)
    and revenue is $0 or negligible, it's likely a sample."""
    if revenue > 10:  # Has real revenue = real sale
        return False
    name = str(name).strip()
    words = name.split()
    if len(words) < 2 or len(words) > 4:
        return False
    # Business keywords that indicate it's NOT a person
    biz_words = {'inc', 'llc', 'corp', 'company', 'co', 'distribut', 'beverage',
                 'wine', 'spirits', 'store', 'market', 'restaurant', 'bar', 'hotel',
                 'group', 'trading', 'associates', 'wholesale', 'cellar', 'shop',
                 'cafe', 'bistro', 'grill', 'pub', 'lounge', 'thorn', 'yolo'}
    name_lower = name.lower()
    for bw in biz_words:
        if bw in name_lower:
            return False
    # All words should be mostly alphabetic (allow for Jr, III, etc.)
    alpha_words = sum(1 for w in words if w.replace('.','').replace(',','').isalpha())
    return alpha_words >= 2


# ============================================================
# EXCEL PARSING
# ============================================================

def load_13w_velocity(filepath):
    """Parse MT_13W_Distributor_SKU_Velocity.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    week_cols_ce = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28]

    # Grand total from row 3
    grand_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    grand = {
        '13w_ce': safe_num(grand_row[30]),
        '13w_rev': safe_num(grand_row[31]),
        'prior_ce': safe_num(grand_row[32]),
        'prior_rev': safe_num(grand_row[33]),
        'weekly_ce': [safe_num(grand_row[c]) for c in week_cols_ce],
    }

    # Week headers from row 1
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    week_headers = []
    for c in week_cols_ce:
        h = row1[c]
        if h:
            week_headers.append(str(h).replace('1 Week ', ''))
        else:
            week_headers.append('')

    vel_dist = {}
    vel_dist_sku = {}

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        item = str(row[1]) if row[1] else ''
        state = str(row[2]) if row[2] else ''
        ym = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if item == 'Total' and state == 'Total' and ym == 'Total':
            weekly = [safe_num(row[c]) for c in week_cols_ce]
            vel_dist[dist] = {
                '13w_ce': safe_num(row[30]),
                '13w_rev': safe_num(row[31]),
                'prior_ce': safe_num(row[32]),
                'prior_rev': safe_num(row[33]),
                'diff_ce': safe_num(row[34]),
                'pct_ce': safe_pct(row[36]),
                'weekly_ce': weekly,
            }
        elif item != 'Total' and state == 'Total' and ym == 'Total':
            wine = map_wine(item)
            weekly = [safe_num(row[c]) for c in week_cols_ce]
            vel_dist_sku[(dist, wine)] = {
                '13w_ce': safe_num(row[30]),
                '13w_rev': safe_num(row[31]),
                'prior_ce': safe_num(row[32]),
                'prior_rev': safe_num(row[33]),
                'weekly_ce': weekly,
            }

    wb.close()
    return grand, vel_dist, vel_dist_sku, week_headers


def load_4w_comparison(filepath):
    """Parse MT_4W_vs_Prior4W_SKU.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    grand_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    four_w_grand = {
        'cur_ce': safe_num(grand_row[3]),
        'prior_ce': safe_num(grand_row[5]),
    }

    four_w_dist = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        item = str(row[2]) if row[2] else ''

        if not dist or dist == 'Total':
            continue
        if state == 'Total' and item == 'Total':
            four_w_dist[dist] = {
                'cur_ce': safe_num(row[3]),
                'prior_ce': safe_num(row[5]),
            }

    wb.close()
    return four_w_grand, four_w_dist


def load_inventory(filepath):
    """Parse MT_Inventory_DaysOnHand.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    inv_dist = {}
    inv_dist_sku = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        item = str(row[2]) if row[2] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and item == 'Total':
            inv_dist[dist] = {
                'on_hand': safe_num(row[3]),
                'daily_rate': safe_num(row[4]),
                'doh': safe_num(row[5]),
                '90d_ce': safe_num(row[6]),
                'proj_order': safe_num(row[11]),
            }
        elif item != 'Total' and state != 'Total':
            wine = map_wine(item)
            inv_dist_sku[(dist, state, wine)] = {
                'on_hand': safe_num(row[3]),
                'daily_rate': safe_num(row[4]),
                'doh': safe_num(row[5]),
            }

    wb.close()
    return inv_dist, inv_dist_sku


def load_placements(base_dir):
    """Parse MT_NewPlacements.xlsx and MT_LostPlacements.xlsx"""
    # New placements
    wb_new = openpyxl.load_workbook(os.path.join(base_dir, 'MT_NewPlacements_30Sold_90Unsold_CaseEq.xlsx'), data_only=True)
    ws_new = wb_new.active

    new_dist = {}
    new_accounts = []
    sample_placements = []  # Track sample entries separately

    for row in ws_new.iter_rows(min_row=3, max_row=ws_new.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        acct = str(row[2]) if row[2] else ''
        item = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and acct == 'Total' and item == 'Total':
            new_dist[dist] = {'sold_ce': safe_num(row[4]), 'new_accts': int(safe_num(row[6]))}
        elif acct != 'Total' and item == 'Total':
            sku_count = 0
            # Count distinct SKUs for this account
            entry = {
                'dist': dist, 'state': state, 'account': acct,
                'sold_ce': safe_num(row[4]), 'new_accts': int(safe_num(row[6])),
            }
            if is_sample_account(acct):
                sample_placements.append(entry)
            else:
                new_accounts.append(entry)
    wb_new.close()

    # Lost placements
    wb_lost = openpyxl.load_workbook(os.path.join(base_dir, 'MT_LostPlacements_30Sold_90Unsold.xlsx'), data_only=True)
    ws_lost = wb_lost.active

    lost_dist = {}
    for row in ws_lost.iter_rows(min_row=4, max_row=ws_lost.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        acct = str(row[2]) if row[2] else ''
        item = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and acct == 'Total' and item == 'Total':
            lost_dist[dist] = {'prior_ce': safe_num(row[5]), 'lost_accts': int(safe_num(row[7]))}
    wb_lost.close()

    return new_dist, new_accounts, lost_dist, sample_placements


def load_rolling_period(filepath):
    """Parse MT_4M_RollingPeriod.xlsx for account-level data, distributor detail, and reorder forecast."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    roll_accounts = {}
    roll_account_skus = {}
    roll_account_sku_detail = {}  # (acct, dist, wine) -> SKU-level per account
    sample_accounts = {}  # Track sample account data separately

    # Distributor+SKU level data for distDetail
    dist_sku_detail = {}  # (dist, wine) -> {ce, rev, prior_ce, ...}

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        item = str(row[0]) if row[0] else ''
        dist = str(row[1]).strip() if row[1] else ''
        oo = str(row[2]).strip() if row[2] else ''
        acct = str(row[3]).strip() if row[3] else ''
        state = str(row[4]).strip() if row[4] else ''

        if not acct or acct == 'Total' or not dist or dist == 'Total':
            continue
        if 'Report Created' in item:
            continue
        if oo == 'Total':
            continue
        if 'YOLO' in acct.upper():
            continue

        if is_sample_account(acct):
            # Track samples separately - don't mix into regular accounts
            skey = (acct, dist)
            if skey not in sample_accounts:
                sample_accounts[skey] = {
                    'account': acct, 'dist': dist, 'state': state, 'channel': oo,
                    'total_ce': 0, 'total_rev': 0,
                }
            sample_accounts[skey]['total_ce'] += safe_num(row[37])
            sample_accounts[skey]['total_rev'] += safe_num(row[38])
            continue  # Skip adding to regular accounts

        wine = map_wine(item)

        # --- Account-level aggregation ---
        key = (acct, dist)
        if key not in roll_accounts:
            roll_accounts[key] = {
                'account': acct, 'dist': dist, 'state': state, 'channel': oo,
                'nov_ce': 0, 'dec_ce': 0, 'jan_ce': 0, 'feb_ce': 0,
                'nov_rev': 0, 'dec_rev': 0, 'jan_rev': 0, 'feb_rev': 0,
                'total_ce': 0, 'total_rev': 0, 'total_purchases': 0,
                'prior_ce': 0, 'prior_rev': 0,
                'first_buy': None, 'last_buy': None,
                'ros_acct': 0, 'eff_placements': 0,
            }
            roll_account_skus[key] = set()

        a = roll_accounts[key]
        roll_account_skus[key].add(wine)

        a['nov_ce'] += safe_num(row[5])
        a['nov_rev'] += safe_num(row[6])
        a['dec_ce'] += safe_num(row[13])
        a['dec_rev'] += safe_num(row[14])
        a['jan_ce'] += safe_num(row[21])
        a['jan_rev'] += safe_num(row[22])
        a['feb_ce'] += safe_num(row[29])
        a['feb_rev'] += safe_num(row[30])
        a['total_ce'] += safe_num(row[37])
        a['total_rev'] += safe_num(row[38])
        a['total_purchases'] += safe_num(row[42])
        a['eff_placements'] += safe_num(row[39])

        ros = safe_num(row[40])
        if ros > a['ros_acct']:
            a['ros_acct'] = ros

        a['prior_ce'] += safe_num(row[45])
        a['prior_rev'] += safe_num(row[46])

        for col in [11, 12, 19, 20, 27, 28, 35, 36, 43, 44]:
            dt = parse_date_str(row[col])
            if dt:
                if not a['first_buy'] or dt < a['first_buy']:
                    a['first_buy'] = dt
                if not a['last_buy'] or dt > a['last_buy']:
                    a['last_buy'] = dt

        # --- Account+SKU level aggregation (for reorder by SKU) ---
        ask = (acct, dist, wine)
        if ask not in roll_account_sku_detail:
            roll_account_sku_detail[ask] = {
                'ce': 0, 'total_purchases': 0,
                'first_buy': None, 'last_buy': None,
                'nov_ce': 0, 'dec_ce': 0, 'jan_ce': 0, 'feb_ce': 0,
            }
        asd = roll_account_sku_detail[ask]
        asd['ce'] += safe_num(row[37])
        asd['total_purchases'] += safe_num(row[42])
        asd['nov_ce'] += safe_num(row[5])
        asd['dec_ce'] += safe_num(row[13])
        asd['jan_ce'] += safe_num(row[21])
        asd['feb_ce'] += safe_num(row[29])

        for bcol in [11, 12, 19, 20, 27, 28, 35, 36, 43, 44]:
            bdt = parse_date_str(row[bcol])
            if bdt:
                if not asd['first_buy'] or bdt < asd['first_buy']:
                    asd['first_buy'] = bdt
                if not asd['last_buy'] or bdt > asd['last_buy']:
                    asd['last_buy'] = bdt

        # --- Distributor+SKU aggregation ---
        dk = (dist, wine)
        if dk not in dist_sku_detail:
            dist_sku_detail[dk] = {
                'ce': 0, 'rev': 0, 'prior_ce': 0, 'net_price': 0, 'net_price_count': 0,
                'weekly_counts': defaultdict(int),
            }
        ds = dist_sku_detail[dk]
        ds['ce'] += safe_num(row[37])
        ds['rev'] += safe_num(row[38])
        ds['prior_ce'] += safe_num(row[45])

        # Net price (column index may vary - check for valid price)
        np_val = safe_num(row[41])
        if np_val > 0:
            ds['net_price'] += np_val
            ds['net_price_count'] += 1

    # Add SKU count
    for key in roll_accounts:
        roll_accounts[key]['sku_count'] = len(roll_account_skus[key])

    wb.close()
    return roll_accounts, roll_account_skus, dist_sku_detail, roll_account_sku_detail, sample_accounts


def match_qb_dist_name(qb_name, dep_dist_names):
    """Fuzzy match a QuickBooks distributor name to a depletion report distributor name."""
    qb_lower = qb_name.lower().strip()
    # Direct keyword matching pairs
    keyword_map = [
        ('shangy', 'shangy'),
        ('fedway', 'fedway'),
        ('pine state', 'pine state'),
        ('quail', 'quail'),
        ('tryon', 'tryon'),
        ('okoboji', 'okoboji'),
        ('libdib', 'libdib'),
        ('winebow', 'winebow'),
        ('craft brewers', 'craft brewers'),
        ('ben arnold', 'ben arnold'),
        ('breakthru beverage nevada', 'breakthru nv'),
        ('premier beverage company (tampa)', 'breakthru fl - riverview'),
        ('premiere beverage company (miramar):connecticut', 'breakthru ct'),
        ('premiere beverage company (miramar)', 'breakthru fl - miramar'),
        ('carisam', 'breakthru fl'),
        ('rndc ga', 'republic national dist - atlanta'),
        ('republic national distributing', 'republic national dist'),
        ('aoc', 'aoc'),
    ]
    for qb_kw, dep_kw in keyword_map:
        if qb_kw in qb_lower:
            matches = [d for d in dep_dist_names if dep_kw in d.lower()]
            if matches:
                return matches  # return all matching depot names
    return []


def classify_qb_product_label(product_name):
    """Classify a QB product name as Classic or Contemporary label.
    Classic: 'Cases of NA Wine:Case of Missing Thorn...' (6-pack, original labels, Batch 1+2)
    Contemporary: '12 Pack' products (12-pack, new labels, Batch 3+)
    Bottles and non-wine items return None.
    """
    p_lower = product_name.lower()
    # Skip non-wine items
    if any(skip in p_lower for skip in ['shipping', 'crv', 'discount', 'tax', 'sample', 'bottle - 750', 'bottle of', 'bottle missing']):
        return None
    if '12 pack' in p_lower or '12pk' in p_lower:
        return 'Contemporary'
    if 'cases of na wine' in p_lower:
        return 'Classic'
    return None


def load_qb_distributor_orders(filepath):
    """Parse QuickBooks Sales by Customer Detail for distributor order history.
    Returns dict: {qb_dist_name: {last_date, last_num, last_total, last_items, order_count, labels: {Classic: qty, Contemporary: qty}}}
    """
    from datetime import datetime
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Collect all distributor transactions grouped by customer + invoice number
    invoices = {}  # (customer, invoice_num) -> {date, items, total}
    for row in ws.iter_rows(min_row=6, values_only=True):
        ctype = str(row[12]).strip() if row[12] else ''
        if ctype != 'Distributors':
            continue
        date_val = row[1]
        num = str(row[3]).strip() if row[3] else ''
        product = str(row[4]).strip() if row[4] else ''
        qty = row[6] if row[6] else 0
        amount = row[8] if row[8] else 0
        cname = str(row[10]).strip() if row[10] else ''
        if not date_val or not cname or cname == 'None' or not num:
            continue
        if isinstance(date_val, datetime):
            dt = date_val
        else:
            try:
                dt = datetime.strptime(str(date_val), '%m/%d/%Y')
            except:
                continue

        key = (cname, num)
        if key not in invoices:
            invoices[key] = {'date': dt, 'items': [], 'total': 0, 'num': num, 'customer': cname}
        invoices[key]['items'].append({'product': product, 'qty': qty})
        invoices[key]['total'] += amount if amount else 0

    # Group by customer, find most recent order + track label types across ALL orders
    dist_orders = {}
    dist_labels = {}  # {cname: {'Classic': total_cases, 'Contemporary': total_cases}}
    for (cname, num), inv in invoices.items():
        if cname not in dist_orders:
            dist_orders[cname] = []
        if cname not in dist_labels:
            dist_labels[cname] = {'Classic': 0, 'Contemporary': 0}

        # Simplify product names and classify labels
        simple_items = []
        for item in inv['items']:
            p = item['product']
            # Classify label type
            label_type = classify_qb_product_label(p)
            if label_type and item['qty'] and item['qty'] > 0:
                dist_labels[cname][label_type] += int(item['qty'])

            # Extract wine type from product name
            wine = 'Other'
            p_lower = p.lower()
            if 'red' in p_lower:
                wine = 'Red'
            elif 'sparkling white' in p_lower or 'sparkling whi' in p_lower:
                wine = 'Spark White'
            elif 'sparkling ros' in p_lower:
                wine = 'Spark Rose'
            elif 'still white' in p_lower or ('white' in p_lower and 'sparkling' not in p_lower):
                wine = 'White'
            elif 'still ros' in p_lower or ('ros' in p_lower and 'sparkling' not in p_lower):
                wine = 'Rose'
            if item['qty'] and item['qty'] > 0:
                simple_items.append({'wine': wine, 'qty': int(item['qty'])})
        dist_orders[cname].append({
            'date': inv['date'],
            'num': inv['num'],
            'items': simple_items,
            'total': inv['total']
        })

    # For each distributor, get the most recent order
    result = {}
    for cname, orders in dist_orders.items():
        orders.sort(key=lambda x: x['date'], reverse=True)
        latest = orders[0]
        # Summarize items
        item_summary = {}
        for it in latest['items']:
            w = it['wine']
            if w not in item_summary:
                item_summary[w] = 0
            item_summary[w] += it['qty']
        summary_str = ', '.join([f"{w} x{q}" for w, q in item_summary.items() if w != 'Other'])

        # Determine label status
        labels = dist_labels.get(cname, {'Classic': 0, 'Contemporary': 0})
        has_classic = labels['Classic'] > 0
        has_contemporary = labels['Contemporary'] > 0
        if has_classic and has_contemporary:
            label_status = 'Both'
        elif has_classic:
            label_status = 'Classic'
        elif has_contemporary:
            label_status = 'Contemporary'
        else:
            label_status = 'Unknown'

        result[cname] = {
            'last_date': latest['date'].strftime('%m/%d/%y'),
            'last_num': latest['num'],
            'last_total': round(latest['total']),
            'last_items': summary_str,
            'order_count': len(orders),
            'label_status': label_status,
            'classic_cases': labels['Classic'],
            'contemporary_cases': labels['Contemporary']
        }

    wb.close()
    return result


def load_warehouse_inventory(filepath):
    """Parse Inventory Tracker Live.xlsx for warehouse stock levels.
    Uses the most recent date-stamped tab for current inventory data.
    Returns dict with classic and contemporary inventory by location and wine type.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    wine_cols = ['Still White', 'Still Rose', 'Red', 'Sparkling White', 'Sparkling Rose']

    # Find most recent date-stamped sheet
    date_sheets = []
    for s in wb.sheetnames:
        if not any(c.isdigit() for c in s) or '.' not in s:
            continue
        if any(kw in s for kw in ('Summary', 'Sales', 'Change', 'Amazon', 'Cases')):
            continue
        parts = s.split('.')
        if len(parts) >= 3:
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                yr = y + 2000 if y < 100 else y
                date_sheets.append((yr, m, d, s))
            except ValueError:
                pass
    if date_sheets:
        date_sheets.sort(reverse=True)
        latest_sheet = date_sheets[0][3]
    else:
        latest_sheet = 'Jan 2026 Summary'

    print(f"  Using warehouse tab: {latest_sheet}")
    ws = wb[latest_sheet]

    def safe_val(v):
        if v is None or str(v).strip() == 'None':
            return 0
        try:
            return round(float(v))
        except:
            return 0

    classic = {}
    contemporary = {}
    batch4 = {}

    # Batch 1+2 (Classic): rows 4-9
    for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
        vals['total'] = safe_val(row[11])
        classic[loc] = vals

    # Batch 3 (Contemporary/New Labels): rows 15-20
    for row in ws.iter_rows(min_row=15, max_row=20, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
        vals['total'] = safe_val(row[11])
        contemporary[loc] = vals

    # Quantity Produced (row 21)
    produced = {}
    for row in ws.iter_rows(min_row=21, max_row=21, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if loc and 'Produced' in loc:
            produced = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            produced['total'] = safe_val(row[11])

    # Grand total and Batch 4 estimates - scan rows 23-27
    grand_total = {}
    for row in ws.iter_rows(min_row=23, max_row=27, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if 'Total' in loc and 'Batch' not in loc and 'Post' not in loc:
            grand_total = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            grand_total['total'] = safe_val(row[11])
        elif 'Batch 4' in loc:
            batch4 = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            batch4['total'] = safe_val(row[11])

    wb.close()
    return {
        'classic': classic,
        'contemporary': contemporary,
        'batch3_produced': produced,
        'batch4_estimates': batch4,
        'grand_total': grand_total,
        'data_date': latest_sheet,
    }


def load_inventory_tracker_live(filepath):
    """Parse Inventory Tracker Live.xlsx for Classic (Batch 1+2) sellout tracker.
    Reads the most recent date-stamped sheet, extracts Batch 1+2 and Batch 3 data.
    Excludes 627 bad Sparkling Rose at Zephyr Batch 1.
    Also reads historical sheets to compute burn rate.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find most recent date sheet by parsing M.D.YY dates and sorting
    date_sheets = []
    for s in wb.sheetnames:
        if not any(c.isdigit() for c in s) or '.' not in s:
            continue
        if any(kw in s for kw in ('Summary', 'Sales', 'Change', 'Amazon', 'Cases')):
            continue
        parts = s.split('.')
        if len(parts) >= 3:
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                yr = y + 2000 if y < 100 else y
                date_sheets.append((yr, m, d, s))
            except ValueError:
                pass
    if date_sheets:
        date_sheets.sort(reverse=True)
        latest_sheet = date_sheets[0][3]
    else:
        latest_sheet = '2.9.26'

    ws = wb[latest_sheet]
    wine_names = ['Still White', 'Still Rose', 'Red', 'Sparkling White', 'Sparkling Rose']

    def sv(v):
        if v is None: return 0
        try: return float(v)
        except: return 0

    # Batch 1+2 (Classic) - Rows 4-9, cols F(6)-L(12)
    classic_locations = {}
    for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {
            'Still White': sv(row[6]), 'Still Rose': sv(row[7]),
            'Red': sv(row[8]), 'Sparkling White': sv(row[9]),
            'Sparkling Rose': sv(row[10]), 'total': sv(row[11])
        }
        classic_locations[loc] = vals

    # Batch 3 (New Labels) - Rows 15-19, cols F-L
    new_label_locations = {}
    for row in ws.iter_rows(min_row=15, max_row=19, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {
            'Still White': sv(row[6]), 'Still Rose': sv(row[7]),
            'Red': sv(row[8]), 'Sparkling White': sv(row[9]),
            'Sparkling Rose': sv(row[10]), 'total': sv(row[11])
        }
        new_label_locations[loc] = vals

    # BAD inventory: 627 Sparkling Rose at Zephyr Express, CA Batch 1
    bad_spr = 0
    for loc, vals in classic_locations.items():
        if 'Batch 1' in loc and 'Zephyr' in loc:
            bad_spr = vals.get('Sparkling Rose', 0)

    # Classic totals (from Warehouse Total row)
    classic_total_row = classic_locations.get('Warehouse Total', {})
    classic_total = classic_total_row.get('total', 0)
    classic_trackable = classic_total - bad_spr

    # Classic by wine (trackable - excluding bad SpR)
    classic_by_wine = {}
    for wine in wine_names:
        classic_by_wine[wine] = classic_total_row.get(wine, 0)
    classic_by_wine['Sparkling Rose'] -= bad_spr

    # New label totals
    new_total_row = new_label_locations.get('Warehouse Total', {})
    new_total = new_total_row.get('total', 0)

    # Historical burn rate from previous sheets (use sorted date_sheets)
    history = []
    historical_sheets = [s[3] for s in sorted(date_sheets)]  # chronological order
    for sname in historical_sheets:
        if sname in wb.sheetnames:
            hws = wb[sname]
            for row in hws.iter_rows(min_row=9, max_row=9, values_only=True):
                loc = str(row[5]).strip() if row[5] else ''
                if 'Warehouse Total' in loc:
                    t = sv(row[11])
                    # Parse date from sheet name
                    parts = sname.replace('.25', '.2025').replace('.26', '.2026').replace('.24', '.2024').split('.')
                    if len(parts) >= 3:
                        date_str = f"{parts[0]}/{parts[1]}/{parts[2]}"
                    elif len(parts) == 2:
                        date_str = sname
                    else:
                        date_str = sname
                    history.append({'date': sname, 'total': t})

    # Compute monthly burn rate from last 3 data points
    burn_rate_monthly = 0
    if len(history) >= 2:
        # Use last 3 months of data for burn rate
        recent = history[-3:] if len(history) >= 3 else history
        total_drop = recent[0]['total'] - recent[-1]['total']
        # Estimate months between data points (roughly 1 month apart)
        n_periods = len(recent) - 1
        if n_periods > 0 and total_drop > 0:
            burn_rate_monthly = total_drop / n_periods

    # Deadline: June 30, 2026
    from datetime import datetime
    today = datetime.now()
    deadline = datetime(2026, 6, 30)
    months_remaining = max(0, (deadline - today).days / 30.44)
    required_monthly = classic_trackable / months_remaining if months_remaining > 0 else 0

    wb.close()

    return {
        'classic_locations': {k: v for k, v in classic_locations.items() if k not in ('Warehouse Total', 'Zephyr Total')},
        'classic_total': classic_total,
        'classic_trackable': round(classic_trackable, 1),
        'classic_by_wine': {k: round(v, 1) for k, v in classic_by_wine.items()},
        'bad_spr': round(bad_spr, 1),
        'new_label_locations': {k: v for k, v in new_label_locations.items() if k != 'Warehouse Total'},
        'new_label_total': round(new_total, 1),
        'burn_rate_monthly': round(burn_rate_monthly, 1),
        'months_remaining': round(months_remaining, 1),
        'required_monthly': round(required_monthly, 1),
        'history': history,
        'data_date': latest_sheet,
    }


def load_quickbooks_revenue(filepath):
    """Parse Sales by Customer Detail - RD.xlsx (QuickBooks) for revenue data.

    File structure (QuickBooks Sales by Customer Detail export):
    Row 1: "Sales by Customer Detail - RD"
    Row 2: "Yolo Brand Group, LLC"
    Row 3: "All Dates"
    Row 4: (blank)
    Row 5: Headers - [blank, Transaction date, Transaction type, Num,
            Product/Service full name, Memo/Description, Quantity,
            Sales price, Amount, Balance, Customer full name,
            Customer company, Customer type]
    Row 6+: Data rows where:
      - Customer name rows: col A has name, rest blank
      - Transaction rows: col A blank, cols B-M have data
      - Total rows: col A = "Total for [customer]"
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    product_rev = defaultdict(lambda: {'rev': 0, 'units': 0})
    case_rev = defaultdict(lambda: {'rev': 0, 'cases': 0})  # Case-only for avg rev/case
    product_rev_2026 = defaultdict(lambda: {'rev': 0, 'units': 0})  # 2026-only SKU data
    case_rev_2026 = defaultdict(lambda: {'rev': 0, 'cases': 0})
    customer_type_rev = defaultdict(float)
    monthly_rev = defaultdict(lambda: defaultdict(float))
    monthly_rev_2026 = defaultdict(lambda: defaultdict(float))  # 2026-only monthly by wine
    monthly_channel_rev = defaultdict(lambda: defaultdict(float))  # 2026 channel x month
    dist_cust_monthly_rev = defaultdict(lambda: defaultdict(float))  # {qb_cust_name: {month: amount}} for distributor txns
    total_rev = 0
    total_units = 0
    total_txns = 0
    total_rev_2026 = 0
    total_units_2026 = 0
    total_txns_2026 = 0

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 9:
            continue

        # Transaction rows have col A (index 0) blank, data in cols B-M (index 1-12)
        # Skip customer name rows (col A has text, rest blank) and Total rows
        col_a = str(row[0]).strip() if row[0] else ''
        if col_a and col_a != 'None':
            continue  # Skip customer header rows and "Total for..." rows

        date_val = row[1]       # Column B: Transaction date
        txn_type = str(row[2]) if row[2] else ''  # Column C: Transaction type
        product = str(row[4]) if row[4] else ''    # Column E: Product/Service full name
        memo = str(row[5]) if row[5] else ''       # Column F: Memo/Description
        qty = safe_num(row[6])                     # Column G: Quantity
        amount = safe_num(row[8])                  # Column I: Amount
        cust_name = str(row[10]) if row[10] else '' # Column K: Customer full name
        cust_type = str(row[12]) if row[12] else '' # Column M: Customer type

        if not date_val or not txn_type:
            continue

        # Skip non-revenue items (shipping, CRV, discounts, etc.)
        product_lower = product.lower()
        if any(skip in product_lower for skip in ['shipping', 'crv', 'discount', 'sales tax', 'tax item']):
            continue

        # Determine wine type from product name and memo/description
        wine = None
        search_text = (product + ' ' + memo).lower()
        if 'still red' in search_text or ('red wine' in search_text and 'sparkling' not in search_text) or 'non alcoholic red' in search_text:
            wine = 'Still Red'
        elif 'still white' in search_text or ('white wine' in search_text and 'sparkling' not in search_text):
            wine = 'Still White'
        elif 'still rose' in search_text or 'still rosé' in search_text or ('rose' in search_text and 'sparkling' not in search_text and 'rosé' not in search_text.split('sparkling')[-1]):
            wine = 'Still Rose'
        elif 'sparkling white' in search_text:
            wine = 'Sparkling White'
        elif 'sparkling rose' in search_text or 'sparkling rosé' in search_text:
            wine = 'Sparkling Rose'

        # Parse date early for year filtering
        dt_parsed = None
        if date_val:
            try:
                if isinstance(date_val, datetime):
                    dt_parsed = date_val
                else:
                    dt_parsed = datetime.strptime(str(date_val).strip(), '%m/%d/%Y')
            except:
                pass

        is_2026 = dt_parsed and dt_parsed.year == 2026

        if amount:
            total_rev += amount
            total_txns += 1
            if is_2026:
                total_rev_2026 += amount
                total_txns_2026 += 1
        if qty:
            total_units += abs(int(qty))
            if is_2026:
                total_units_2026 += abs(int(qty))

        if wine and amount:
            product_rev[wine]['rev'] += amount
            product_rev[wine]['units'] += abs(int(qty)) if qty else 0

            # 2026-only product tracking
            if is_2026:
                product_rev_2026[wine]['rev'] += amount
                product_rev_2026[wine]['units'] += abs(int(qty)) if qty else 0

            # Track case-only revenue for accurate avg rev/case
            if amount > 0 and qty and qty > 0:
                prod_lower = product_lower
                if 'case of missing thorn' in prod_lower or 'case missing thorn' in prod_lower:
                    case_rev[wine]['rev'] += amount
                    case_rev[wine]['cases'] += qty
                    if is_2026:
                        case_rev_2026[wine]['rev'] += amount
                        case_rev_2026[wine]['cases'] += qty

        # Classify by customer type (from Column M), with fallback by customer name
        if amount:
            ct = cust_type.strip()
            # If no customer type, infer from customer name
            if not ct:
                cust_lower = cust_name.lower().strip()
                # Customer name -> channel override map
                CUSTOMER_CHANNEL_MAP = {
                    'qvc': 'Direct to Trade - Off Premise',
                    "fleming's": 'Direct to Trade - On Premise',
                }
                for keyword, channel in CUSTOMER_CHANNEL_MAP.items():
                    if keyword in cust_lower:
                        ct = channel
                        break

            if ct:
                # Map QuickBooks customer types to dashboard categories
                ct_lower = ct.lower()
                if 'distributor' in ct_lower:
                    customer_type_rev['Distributors'] += amount
                elif 'website' in ct_lower or 'dtc' in ct_lower or 'fbm' in ct_lower:
                    customer_type_rev['Website / DTC (FBM)'] += amount
                elif 'on premise' in ct_lower:
                    customer_type_rev['Direct to Trade - On Premise'] += amount
                elif 'off premise' in ct_lower:
                    customer_type_rev['Direct to Trade - Off Premise'] += amount
                elif 'promo' in ct_lower:
                    customer_type_rev['Promo'] += amount
                elif 'operations' in ct_lower:
                    customer_type_rev['Operations'] += amount
                else:
                    customer_type_rev[ct] += amount
            else:
                customer_type_rev['Uncategorized'] += amount

        # Monthly revenue by wine
        if wine and amount and dt_parsed:
            month_key = dt_parsed.strftime('%Y-%m')
            monthly_rev[wine][month_key] += amount
            if is_2026:
                monthly_rev_2026[wine][month_key] += amount

        # 2026 monthly revenue by channel type
        if amount and is_2026:
            ct = cust_type.strip() if cust_type else ''
            # Infer channel from customer name if type is blank
            if not ct:
                cust_lower = cust_name.lower().strip()
                CUSTOMER_CHANNEL_MAP_2026 = {
                    'qvc': 'Direct to Trade - Off Premise',
                    "fleming's": 'Direct to Trade - On Premise',
                }
                for keyword, ch_override in CUSTOMER_CHANNEL_MAP_2026.items():
                    if keyword in cust_lower:
                        ct = ch_override
                        break
            ct_lower = ct.lower()
            if 'distributor' in ct_lower:
                channel = 'Distributors'
            elif 'website' in ct_lower or 'dtc' in ct_lower or 'fbm' in ct_lower:
                channel = 'Website / DTC'
            elif 'on premise' in ct_lower:
                channel = 'Direct to Trade - On Premise'
            elif 'off premise' in ct_lower:
                channel = 'Direct to Trade - Off Premise'
            elif 'promo' in ct_lower:
                channel = 'Promo'
            elif 'operations' in ct_lower:
                channel = 'Operations'
            else:
                channel = 'Other'
            month_key = dt_parsed.strftime('%Y-%m')
            monthly_channel_rev[channel][month_key] += amount
            # Track per-customer distributor revenue for state filtering
            if channel == 'Distributors' and cust_name:
                dist_cust_monthly_rev[cust_name.strip()][month_key] += amount

    wb.close()

    # Build product mix (use case-only data for avg rev/case)
    product_mix = []
    for wine in WINE_ORDER:
        if wine in product_rev:
            p = product_rev[wine]
            cr = case_rev.get(wine, {'rev': 0, 'cases': 0})
            avg_case = round(cr['rev'] / cr['cases'], 2) if cr['cases'] > 0 else 0
            product_mix.append({
                'name': wine,
                'rev': round(p['rev'], 2),
                'units': p['units'],
                'avgCase': avg_case,
            })

    # Build 2026-only product mix
    product_mix_2026 = []
    for wine in WINE_ORDER:
        if wine in product_rev_2026:
            p = product_rev_2026[wine]
            cr = case_rev_2026.get(wine, {'rev': 0, 'cases': 0})
            avg_case = round(cr['rev'] / cr['cases'], 2) if cr['cases'] > 0 else 0
            product_mix_2026.append({
                'name': wine,
                'rev': round(p['rev'], 2),
                'units': p['units'],
                'avgCase': avg_case,
            })

    # Build customer revenue
    cust_rev = [{'type': k, 'rev': round(v, 2)} for k, v in sorted(customer_type_rev.items(), key=lambda x: -x[1])]

    # Build 2026 monthly channel revenue sorted output
    channel_ytd = {}
    for ch, months in monthly_channel_rev.items():
        channel_ytd[ch] = {m: round(v, 2) for m, v in sorted(months.items())}

    # Build per-customer distributor YTD rev
    dist_cust_ytd = {}
    for cname, months in dist_cust_monthly_rev.items():
        dist_cust_ytd[cname] = {m: round(v, 2) for m, v in sorted(months.items())}

    return (product_mix, cust_rev, monthly_rev, round(total_rev, 2), total_units, total_txns,
            channel_ytd, product_mix_2026, monthly_rev_2026,
            round(total_rev_2026, 2), total_units_2026, total_txns_2026, dist_cust_ytd)


# ============================================================
# DATA TRANSFORMATION
# ============================================================

def compute_inventory_status(doh):
    """Classify inventory health based on Days on Hand."""
    if doh <= 0:
        return 'Critical'
    if doh <= 90:
        return 'Low'
    if doh <= 180:
        return 'Healthy'
    if doh <= 365:
        return 'Overstocked'
    return 'Dead Stock'  # 365+ days


def build_dist_scorecard(vel_dist, four_w_dist, inv_dist, new_dist, lost_dist):
    """Build the distScorecard array for the dashboard."""
    scorecard = []
    for dist, v in vel_dist.items():
        # Get state from distributor name
        st = get_state_from_dist(dist)

        # 4W data
        fw = four_w_dist.get(dist, {})
        w4_ce = fw.get('cur_ce', 0)
        prior_w4 = fw.get('prior_ce', 0)

        # Inventory
        inv = inv_dist.get(dist, {})
        oh = inv.get('on_hand', 0)
        doh = inv.get('doh', 0)

        # Placements
        nd = new_dist.get(dist, {})
        ld = lost_dist.get(dist, {})
        net = nd.get('new_accts', 0) - ld.get('lost_accts', 0)

        # Consistency (weeks with activity / 13)
        weekly = v.get('weekly_ce', [0]*13)
        active_weeks = sum(1 for w in weekly if w > 0)
        con = round(active_weeks / 13, 2)

        # YoY
        yoy = fmt_yoy(v.get('prior_ce', 0), v.get('13w_ce', 0))
        mom = fmt_mom(prior_w4, w4_ce)

        # Market stage classification
        weekly = v.get('weekly_ce', [0]*13)
        if v.get('prior_ce', 0) < 2.0:
            stage = 'Launch'
        elif con >= 0.7 and v.get('13w_ce', 0) > 20:
            stage = 'Mature'
        else:
            stage = 'Growth'

        # Health score (0-100): weighted composite
        # Velocity component (30%): 4W vs 13W avg pace
        avg_weekly = v.get('13w_ce', 0) / 13
        vel_score = min(100, (w4_ce / (avg_weekly * 4) * 100)) if avg_weekly > 0 else 0
        # Inventory component (25%): penalize <60 or >300 DOH
        capped_doh = min(doh, 999)  # Cap extreme DOH for scoring
        inv_score = 100 if 60 <= capped_doh <= 300 else max(0, 100 - abs(capped_doh - 180) * 0.5) if capped_doh > 0 else 0
        # Placement component (20%): net placements
        plc_score = min(100, max(0, 50 + net * 10))
        # Consistency component (25%)
        con_score = con * 100
        health = round(vel_score * 0.30 + inv_score * 0.25 + plc_score * 0.20 + con_score * 0.25)

        # Velocity trend: compare last 4 weeks vs first 4 weeks of 13W
        first4 = sum(weekly[:4])
        last4 = sum(weekly[-4:])
        if first4 < 0.5:  # No meaningful prior data
            vel_trend = 'new'
        elif last4 >= first4 * 1.1:
            vel_trend = 'up'
        elif last4 <= first4 * 0.9:
            vel_trend = 'down'
        else:
            vel_trend = 'flat'

        # Sell-through rate: 13W CE / (13W CE + current OH) * 100
        total_product = v.get('13w_ce', 0) + oh
        sell_through = round((v.get('13w_ce', 0) / total_product) * 100, 1) if total_product > 0 else 0

        scorecard.append({
            'name': dist, 'st': st,
            'ce': round(v.get('13w_ce', 0), 1),
            'rev': round(v.get('13w_rev', 0)),
            'prior': round(v.get('prior_ce', 0), 1),
            'yoy': yoy,
            'w4': round(w4_ce, 1),
            'prior_w4': round(prior_w4, 1),
            'mom': mom,
            'oh': round(oh, 1),
            'doh': round(doh),
            'net': net,
            'con': con,
            'weekly': [round(w, 1) for w in weekly],
            'stage': stage,
            'health': health,
            'velTrend': vel_trend,
            'sellThru': sell_through,
        })

    # Sort by 13W CE descending
    scorecard.sort(key=lambda x: -x['ce'])
    return scorecard


def build_inventory_data(inv_dist, inv_dist_sku=None):
    """Build the inventoryData array for the dashboard, including SKU breakout."""
    inv_data = []
    for dist, inv in inv_dist.items():
        st = get_state_from_dist(dist)
        doh = inv.get('doh', 0)
        status = compute_inventory_status(doh)

        # Build SKU breakout if available
        skus = []
        if inv_dist_sku:
            for wine in WINE_ORDER:
                for (d, s, w), idata in inv_dist_sku.items():
                    if d == dist and w == wine:
                        sku_oh = round(idata.get('on_hand', 0), 1)
                        sku_rate = round(idata.get('daily_rate', 0), 2)
                        sku_doh = round(idata.get('doh', 0))
                        if sku_oh > 0 or sku_rate > 0:
                            skus.append({
                                'w': wine,
                                'oh': sku_oh,
                                'rate': sku_rate,
                                'doh': sku_doh,
                                'status': compute_inventory_status(sku_doh),
                            })
                        break

        inv_data.append({
            'name': dist, 'st': st,
            'oh': round(inv.get('on_hand', 0), 1),
            'rate': round(inv.get('daily_rate', 0), 2),
            'doh': round(doh),
            'dep90': round(inv.get('90d_ce', 0), 1),
            'status': status,
            'proj': round(inv.get('proj_order', 0), 1),
            'skus': skus,
        })

    inv_data.sort(key=lambda x: -x['oh'])
    return inv_data


def build_placement_data(new_dist, lost_dist):
    """Build placementData array (net new/lost by distributor)."""
    all_dists = set(list(new_dist.keys()) + list(lost_dist.keys()))
    placements = []
    for dist in all_dists:
        nd = new_dist.get(dist, {})
        ld = lost_dist.get(dist, {})
        new_a = nd.get('new_accts', 0)
        lost_a = ld.get('lost_accts', 0)
        net = new_a - lost_a
        st = get_state_from_dist(dist)
        short = get_dist_short(dist)

        placements.append({
            'name': short, 'st': st,
            'fullName': dist,  # Keep full name for group matching
            'newA': new_a, 'lostA': lost_a, 'net': net,
        })

    placements.sort(key=lambda x: x['net'])
    return placements


# ============================================================
# MULTI-WAREHOUSE CONSOLIDATION
# ============================================================

def consolidate_scorecard(scorecard):
    """Consolidate multi-warehouse distributors into parent groups with children."""
    groups = {}      # parent_name -> list of entries
    singles = []

    for entry in scorecard:
        grp = get_warehouse_group(entry['name'])
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)

    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        # Sum numeric fields
        ce = sum(c['ce'] for c in children)
        rev = sum(c['rev'] for c in children)
        prior = sum(c['prior'] for c in children)
        w4 = sum(c['w4'] for c in children)
        prior_w4 = sum(c.get('prior_w4', 0) for c in children)
        oh = sum(c['oh'] for c in children)
        net = sum(c['net'] for c in children)

        # DOH: total_oh / total_daily_rate
        total_rate = sum(c['oh'] / c['doh'] if c['doh'] > 0 else 0 for c in children)
        doh = oh / total_rate if total_rate > 0 else 0

        # Consistency: max across children
        con = max(c['con'] for c in children)

        # Weekly: sum across children
        weekly = [0.0] * 13
        for c in children:
            for i, w in enumerate(c.get('weekly', [0]*13)):
                weekly[i] += w

        # Recalculate YoY and MoM from summed values
        yoy = fmt_yoy(prior, ce)
        mom = fmt_mom(prior_w4, w4)

        # Stage and health for grouped entries
        stage = 'Growth'  # Default for combined groups
        # Health: average the children's health scores
        health = round(sum(c['health'] for c in children) / len(children)) if children else 0

        # Velocity trend for group: based on summed weekly
        first4 = sum(weekly[:4])
        last4 = sum(weekly[-4:])
        if first4 < 0.5:
            vel_trend = 'new'
        elif last4 >= first4 * 1.1:
            vel_trend = 'up'
        elif last4 <= first4 * 0.9:
            vel_trend = 'down'
        else:
            vel_trend = 'flat'

        # Sell-through for group
        total_product = ce + oh
        sell_through = round((ce / total_product) * 100, 1) if total_product > 0 else 0

        result.append({
            'name': grp_name, 'st': cfg['st'],
            'ce': round(ce, 1), 'rev': round(rev),
            'prior': round(prior, 1), 'yoy': yoy,
            'w4': round(w4, 1), 'prior_w4': round(prior_w4, 1), 'mom': mom,
            'oh': round(oh, 1), 'doh': round(doh),
            'net': net, 'con': round(con, 2),
            'weekly': [round(w, 1) for w in weekly],
            'isGroup': True,
            'stage': stage,
            'health': health,
            'velTrend': vel_trend,
            'sellThru': sell_through,
            'children': sorted(children, key=lambda x: -x['ce']),
        })

    result.extend(singles)
    result.sort(key=lambda x: -x['ce'])
    return result


def consolidate_inventory(inv_data):
    """Consolidate multi-warehouse inventory into parent groups with children."""
    groups = {}
    singles = []

    for entry in inv_data:
        grp = get_warehouse_group(entry['name'])
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)

    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        oh = sum(c['oh'] for c in children)
        rate = sum(c['rate'] for c in children)
        dep90 = sum(c['dep90'] for c in children)
        proj = sum(c['proj'] for c in children)
        doh = oh / rate if rate > 0 else 0
        status = compute_inventory_status(doh)

        # Aggregate SKUs across warehouses
        sku_agg = {}
        for c in children:
            for s in c.get('skus', []):
                if s['w'] not in sku_agg:
                    sku_agg[s['w']] = {'oh': 0, 'rate': 0}
                sku_agg[s['w']]['oh'] += s['oh']
                sku_agg[s['w']]['rate'] += s['rate']

        consolidated_skus = []
        for wine in WINE_ORDER:
            if wine in sku_agg:
                a = sku_agg[wine]
                s_doh = a['oh'] / a['rate'] if a['rate'] > 0 else 0
                consolidated_skus.append({
                    'w': wine, 'oh': round(a['oh'], 1),
                    'rate': round(a['rate'], 2), 'doh': round(s_doh),
                    'status': compute_inventory_status(s_doh),
                })

        result.append({
            'name': grp_name, 'st': cfg['st'],
            'oh': round(oh, 1), 'rate': round(rate, 2),
            'doh': round(doh), 'dep90': round(dep90, 1),
            'status': status, 'proj': round(proj, 1),
            'skus': consolidated_skus,
            'isGroup': True,
            'children': sorted(children, key=lambda x: -x['oh']),
        })

    result.extend(singles)
    result.sort(key=lambda x: -x['oh'])
    return result


def consolidate_placements(placements):
    """Consolidate multi-warehouse placements into parent groups."""
    groups = {}
    singles = []

    for entry in placements:
        grp = get_warehouse_group(entry.get('fullName', ''))
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)

    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        result.append({
            'name': grp_name, 'st': cfg['st'],
            'newA': sum(c['newA'] for c in children),
            'lostA': sum(c['lostA'] for c in children),
            'net': sum(c['net'] for c in children),
            'isGroup': True,
            'children': children,
        })

    result.extend(singles)
    result.sort(key=lambda x: x['net'])
    return result


def build_new_accounts_list(new_accounts_raw):
    """Build top new accounts by CE sold."""
    # Aggregate by account
    acct_agg = {}
    for a in new_accounts_raw:
        key = a['account']
        if key not in acct_agg:
            acct_agg[key] = {'acct': a['account'], 'dist': a['dist'], 'st': a['state'], 'ce': 0, 'skus': 0}
        acct_agg[key]['ce'] += a['sold_ce']
        acct_agg[key]['skus'] += a.get('new_accts', 1)

    all_accts = sorted(acct_agg.values(), key=lambda x: -x['ce'])
    return [{'acct': a['acct'], 'dist': a['dist'], 'st': a['st'],
             'ce': round(a['ce'], 1), 'skus': a['skus']} for a in all_accts]


def build_accounts_top(roll_accounts, roll_account_skus):
    """Build top 20 accounts by 4M CE."""
    accounts = []
    for key, a in roll_accounts.items():
        if a['total_ce'] <= 0:
            continue

        # Determine trend
        jan = a['jan_ce']
        feb = a['feb_ce']
        nov = a['nov_ce']
        dec = a['dec_ce']
        first_half = nov + dec
        second_half = jan + feb

        if first_half == 0 and second_half > 0:
            trend = 'Accelerating'
        elif second_half > first_half * 1.1:
            trend = 'Accelerating'
        elif second_half < first_half * 0.7:
            trend = 'Decelerating'
        else:
            trend = 'Steady'

        ch = a['channel'] if a['channel'] in ('ON', 'OFF') else 'OFF'

        accounts.append({
            'rank': 0,
            'acct': a['account'],
            'dist': a['dist'],
            'st': a['state'],
            'ch': ch,
            'nov': round(a['nov_ce'], 1),
            'dec': round(a['dec_ce'], 1),
            'jan': round(a['jan_ce'], 1),
            'feb': round(a['feb_ce'], 1),
            'total': round(a['total_ce'], 1),
            'rev': round(a['total_rev']),
            'trend': trend,
        })

    accounts.sort(key=lambda x: -x['total'])
    for i, a in enumerate(accounts):
        a['rank'] = i + 1

    return accounts  # Return ALL accounts; JS will handle show top/all toggle


def build_reorder_data(roll_accounts, roll_account_sku_detail=None):
    """Build reorder forecast with SKU-level breakout from rolling account data."""
    today = datetime.now()
    reorder = []

    for key, a in roll_accounts.items():
        if a['total_ce'] <= 0 or not a['last_buy']:
            continue

        days_since = (today - a['last_buy']).days
        purchases = int(a['total_purchases'])

        if purchases > 1 and a['first_buy'] and a['last_buy'] and a['first_buy'] != a['last_buy']:
            span = (a['last_buy'] - a['first_buy']).days
            avg_cycle = span / (purchases - 1)
        else:
            avg_cycle = 30

        ch = a['channel'] if a['channel'] in ('ON', 'OFF') else 'OFF'

        # Build SKU-level breakout
        skus = []
        if roll_account_sku_detail:
            acct, dist = key
            for wine in WINE_ORDER:
                ask = (acct, dist, wine)
                asd = roll_account_sku_detail.get(ask)
                if not asd or asd['ce'] <= 0:
                    continue

                sku_purch = int(asd['total_purchases'])
                sku_days = 0
                sku_cycle = 0
                sku_last = ''

                if asd['last_buy']:
                    sku_days = (today - asd['last_buy']).days
                    sku_last = asd['last_buy'].strftime('%m/%d/%y')
                    if sku_purch > 1 and asd['first_buy'] and asd['first_buy'] != asd['last_buy']:
                        sku_span = (asd['last_buy'] - asd['first_buy']).days
                        sku_cycle = round(sku_span / (sku_purch - 1))
                    else:
                        sku_cycle = round(avg_cycle)
                else:
                    sku_cycle = round(avg_cycle)

                months_active = sum(1 for m in [asd['nov_ce'], asd['dec_ce'], asd['jan_ce'], asd['feb_ce']] if m > 0)

                skus.append({
                    'w': wine,
                    'ce': round(asd['ce'], 1),
                    'purch': sku_purch,
                    'cycle': sku_cycle,
                    'last': sku_last,
                    'days': sku_days,
                    'months': months_active,
                })

        reorder.append({
            'rank': 0,
            'acct': a['account'],
            'dist': a['dist'],
            'st': a['state'],
            'ch': ch,
            'ce': round(a['total_ce'], 1),
            'purch': purchases,
            'cycle': round(avg_cycle),
            'last': a['last_buy'].strftime('%m/%d/%y'),
            'days': days_since,
            'skus': skus,
        })

    # Sort by CE descending; return ALL - JS will handle show top/all toggle
    reorder.sort(key=lambda x: -x['ce'])
    for i, r in enumerate(reorder):
        r['rank'] = i + 1

    return reorder


def build_dist_detail(vel_dist, vel_dist_sku, four_w_dist, dist_sku_detail):
    """Build the distDetail array with per-SKU breakdown."""
    detail = []

    for dist in sorted(vel_dist.keys(), key=lambda d: -vel_dist[d].get('13w_ce', 0)):
        st = get_state_from_dist(dist)
        # Check if this distributor has unreliable pricing
        # Must match BOTH a pattern (e.g. "Republic") AND a location (e.g. "Dallas", "Houston")
        dist_full = f'{dist}, {st}' if st else dist
        no_price = False
        for pattern in NO_PRICE_PATTERNS:
            if pattern.lower() in dist_full.lower():
                for loc in NO_PRICE_DISTRIBUTORS:
                    if loc.lower() in dist_full.lower():
                        no_price = True
                        break

        wines = []
        for wine in WINE_ORDER:
            sku = vel_dist_sku.get((dist, wine), {})
            ce = round(sku.get('13w_ce', 0), 1)
            if ce <= 0:
                continue

            rev = round(sku.get('13w_rev', 0))
            prior = round(sku.get('prior_ce', 0), 1)
            yoy = fmt_yoy(prior, ce)

            # 4W CE for this SKU (approximate from ratio)
            v = vel_dist.get(dist, {})
            fw = four_w_dist.get(dist, {})
            total_13w = v.get('13w_ce', 1)
            w4_total = fw.get('cur_ce', 0)
            w4 = round(w4_total * (ce / total_13w), 1) if total_13w > 0 else 0

            avg = round(ce / 13, 2)
            rpc = round(rev / ce, 2) if ce > 0 else 0

            # Consistency
            sku_weekly = sku.get('weekly_ce', [0]*13)
            con = round(sum(1 for w in sku_weekly if w > 0) / 13, 2)

            if no_price:
                rev = 0
                rpc = 0

            wines.append({
                'w': wine, 'ce': ce, 'rev': rev, 'prior': prior, 'yoy': yoy,
                'w4': w4, 'avg': avg, 'rpc': rpc, 'con': con,
            })

        if wines:
            entry = {'dist': f'{dist}, {st}' if st and f', {st}' not in dist else dist, 'st': st, 'wines': wines}
            if no_price:
                entry['noPrice'] = True
            detail.append(entry)

    return detail


def build_rev_trend(monthly_rev):
    """Build monthly revenue trend arrays from QuickBooks data."""
    # Get all months across all wines
    all_months = set()
    for wine, months in monthly_rev.items():
        all_months.update(months.keys())

    if not all_months:
        return [], [], {}, []

    sorted_months = sorted(all_months)

    # Generate labels
    labels = []
    for m in sorted_months:
        dt = datetime.strptime(m + '-01', '%Y-%m-%d')
        if dt.month == 1 or dt.month == 6 or len(labels) == 0:
            labels.append(dt.strftime('%b %y'))
        else:
            labels.append(dt.strftime('%b'))

    # Build trend data
    trend = {}
    for wine in WINE_ORDER:
        trend[wine] = [round(monthly_rev.get(wine, {}).get(m, 0), 2) for m in sorted_months]

    # Total by month
    totals = []
    for i, m in enumerate(sorted_months):
        total = sum(trend.get(w, [0]*(i+1))[i] for w in WINE_ORDER)
        totals.append(round(total, 2))

    return sorted_months, labels, trend, totals


# ============================================================
# JAVASCRIPT DATA GENERATION
# ============================================================

def generate_js_data(scorecard, inventory, placements, new_accts, accounts_top,
                     reorder, dist_detail, product_mix, cust_rev,
                     rev_months, rev_labels, rev_trend, rev_totals,
                     rev_total, rev_units, rev_txns, week_headers=None,
                     qb_dist_orders=None, warehouse_inv=None,
                     channel_ytd=None, product_mix_2026=None,
                     rev_months_2026=None, rev_labels_2026=None,
                     rev_trend_2026=None, rev_total_2026=0,
                     rev_units_2026=0, rev_txns_2026=0,
                     classic_tracker=None, sample_placements=None, sample_accounts=None,
                     channel_ytd_by_state=None):
    """Generate the JavaScript data section to inject into HTML."""

    lines = []

    # State names
    states_found = set()
    for s in scorecard:
        if s['st']:
            states_found.add(s['st'])
    for s in inventory:
        if s['st']:
            states_found.add(s['st'])

    state_entries = ','.join(f"{st}:'{STATE_NAMES.get(st, st)}'" for st in sorted(states_found) if st in STATE_NAMES)
    lines.append(f'const stateNames = {{\n  {state_entries}\n}};')
    lines.append('')

    lines.append(f'const regionMap = {json.dumps(REGION_MAP)};')
    lines.append('')

    # ===== DEPLETION DATA =====
    lines.append('// ==================== DEPLETION DATA ====================')
    def _js_scorecard_entry(s, indent='  '):
        weekly_str = ','.join(str(w) for w in s['weekly'])
        base = f"{indent}{{name:'{js_str(s['name'])}',st:'{s['st']}',ce:{s['ce']},rev:{s['rev']},prior:{s['prior']},yoy:'{s['yoy']}',w4:{s['w4']},mom:'{s['mom']}',oh:{s['oh']},doh:{s['doh']},net:{s['net']},con:{s['con']},weekly:[{weekly_str}],stage:'{s['stage']}',health:{s['health']},velTrend:'{s.get('velTrend','flat')}',sellThru:{s.get('sellThru', 0)}"
        if s.get('isGroup') and s.get('children'):
            kids = []
            for c in s['children']:
                kids.append(_js_scorecard_entry(c, indent='    '))
            base += f",isGroup:true,children:[\n" + ',\n'.join(kids) + f"\n{indent}]"
        base += '}'
        return base

    lines.append('const distScorecard = [')
    for s in scorecard:
        lines.append(_js_scorecard_entry(s) + ',')
    lines.append('];')
    lines.append('')

    def _js_inv_entry(inv, indent='  '):
        sku_parts = []
        for s in inv.get('skus', []):
            sku_parts.append(f"{{w:'{s['w']}',oh:{s['oh']},rate:{s['rate']},doh:{s['doh']},status:'{s['status']}'}}")
        skus_str = ','.join(sku_parts)
        base = f"{indent}{{name:'{js_str(inv['name'])}',st:'{inv['st']}',oh:{inv['oh']},rate:{inv['rate']},doh:{inv['doh']},dep90:{inv['dep90']},status:'{inv['status']}',proj:{inv['proj']},skus:[{skus_str}]"
        if inv.get('isGroup') and inv.get('children'):
            kids = []
            for c in inv['children']:
                kids.append(_js_inv_entry(c, indent='    '))
            base += f",isGroup:true,children:[\n" + ',\n'.join(kids) + f"\n{indent}]"
        base += '}'
        return base

    # ===== INVENTORY DATA =====
    lines.append('const inventoryData = [')
    for inv in inventory:
        lines.append(_js_inv_entry(inv) + ',')
    lines.append('];')
    lines.append('')

    # ===== PLACEMENT DATA =====
    lines.append('const placementData = [')
    for p in placements:
        base = f"  {{name:'{js_str(p['name'])}',st:'{p['st']}',newA:{p['newA']},lostA:{p['lostA']},net:{p['net']}"
        if p.get('isGroup') and p.get('children'):
            kids = ','.join(f"{{name:'{js_str(c['name'])}',st:'{c['st']}',newA:{c['newA']},lostA:{c['lostA']},net:{c['net']}}}" for c in p['children'])
            base += f",isGroup:true,children:[{kids}]"
        lines.append(base + '},')
    lines.append('];')
    lines.append('')

    # ===== NEW ACCOUNTS =====
    lines.append('const newAccounts = [')
    for a in new_accts:
        lines.append(f"  {{acct:'{js_str(a['acct'])}',dist:'{js_str(a['dist'])}',st:'{a['st']}',ce:{a['ce']},skus:{a['skus']}}},")
    lines.append('];')
    lines.append('')

    # ===== ACCOUNTS TOP =====
    lines.append('const accountsTop = [')
    for a in accounts_top:
        lines.append(f"  {{rank:{a['rank']},acct:'{js_str(a['acct'])}',dist:'{js_str(a['dist'])}',st:'{a['st']}',ch:'{a['ch']}',nov:{a['nov']},dec:{a['dec']},jan:{a['jan']},feb:{a['feb']},total:{a['total']},rev:{a['rev']},trend:'{a['trend']}'}},")
    lines.append('];')
    lines.append('')

    # ===== ACCOUNT CONCENTRATION =====
    lines.append('// ==================== ACCOUNT CONCENTRATION ====================')
    # Calculate concentration metrics from accountsTop
    if accounts_top:
        all_totals = sorted([a['total'] for a in accounts_top], reverse=True)
        grand_total = sum(all_totals)
        if grand_total > 0:
            top5_pct = round(sum(all_totals[:5]) / grand_total * 100, 1) if len(all_totals) >= 5 else 0
            top10_pct = round(sum(all_totals[:10]) / grand_total * 100, 1) if len(all_totals) >= 10 else 0
            top20_pct = round(sum(all_totals[:20]) / grand_total * 100, 1) if len(all_totals) >= 20 else 0
            median_ce = round(all_totals[len(all_totals)//2], 1)
            under1ce = sum(1 for t in all_totals if t < 1.0)
        else:
            top5_pct = top10_pct = top20_pct = median_ce = 0
            under1ce = 0
        lines.append(f'const acctConcentration = {{total:{len(all_totals)},top5:{top5_pct},top10:{top10_pct},top20:{top20_pct},median:{median_ce},under1:{under1ce}}};')
    else:
        lines.append('const acctConcentration = {total:0,top5:0,top10:0,top20:0,median:0,under1:0};')
    lines.append('')

    # ===== SAMPLE / MARKETING DATA =====
    lines.append('// ==================== SAMPLE / MARKETING DATA ====================')
    if sample_placements is None:
        sample_placements = []
    if sample_accounts is None:
        sample_accounts = {}

    sample_by_dist = defaultdict(lambda: {'ce': 0, 'rev': 0, 'count': 0})
    for (acct, dist), sdata in sample_accounts.items():
        sample_by_dist[dist]['ce'] += sdata['total_ce']
        sample_by_dist[dist]['rev'] += sdata['total_rev']
        sample_by_dist[dist]['count'] += 1
    # Add sample placements
    for sp in sample_placements:
        sample_by_dist[sp['dist']]['placement_ce'] = sample_by_dist[sp['dist']].get('placement_ce', 0) + sp['sold_ce']
        sample_by_dist[sp['dist']]['placement_count'] = sample_by_dist[sp['dist']].get('placement_count', 0) + sp['new_accts']

    sample_list = []
    for dist, sd in sorted(sample_by_dist.items(), key=lambda x: -x[1]['ce']):
        st = ''
        for entry in scorecard:
            if entry['name'] == dist:
                st = entry.get('st', '')
                break
        sample_list.append(f"{{dist:'{js_str(dist)}',st:'{st}',ce:{round(sd['ce'],1)},rev:{round(sd.get('rev',0),2)},accts:{sd['count']},placeCE:{round(sd.get('placement_ce',0),1)},placeCount:{sd.get('placement_count',0)}}}")
    lines.append(f"const sampleSummary = [{','.join(sample_list)}];")
    lines.append('')

    # ===== REORDER DATA (with SKU breakout) =====
    lines.append('const reorderData = [')
    for r in reorder:
        sku_parts = []
        for s in r.get('skus', []):
            sku_parts.append(f"{{w:'{s['w']}',ce:{s['ce']},purch:{s['purch']},cycle:{s['cycle']},last:'{s['last']}',days:{s['days']},months:{s['months']}}}")
        skus_str = ','.join(sku_parts)
        lines.append(f"  {{rank:{r['rank']},acct:'{js_str(r['acct'])}',dist:'{js_str(r['dist'])}',st:'{r['st']}',ch:'{r['ch']}',ce:{r['ce']},purch:{r['purch']},cycle:{r['cycle']},last:'{r['last']}',days:{r['days']},skus:[{skus_str}]}},")
    lines.append('];')
    lines.append('')

    # ===== DISTRIBUTOR DETAIL =====
    lines.append('const distDetail = [')
    for d in dist_detail:
        np_str = "noPrice:true," if d.get('noPrice') else ""
        lines.append(f"  {{dist:'{js_str(d['dist'])}',st:'{d['st']}',{np_str}wines:[")
        for w in d['wines']:
            lines.append(f"    {{w:'{w['w']}',ce:{w['ce']},rev:{w['rev']},prior:{w['prior']},yoy:'{w['yoy']}',w4:{w['w4']},avg:{w['avg']},rpc:{w['rpc']},con:{w['con']}}},")
        lines.append('  ]},')
    lines.append('];')
    lines.append('')

    # ===== REVENUE DATA =====
    lines.append("// ==================== REVENUE DATA (from QuickBooks - company-wide) ====================")
    lines.append('const productMix = [')
    for p in product_mix:
        lines.append(f"  {{name:'{p['name']}',rev:{p['rev']},units:{p['units']},avgCase:{p['avgCase']}}},")
    lines.append('];')
    lines.append('')

    lines.append('const customerRevenue = [')
    for c in cust_rev:
        lines.append(f"  {{type:'{js_str(c['type'])}',rev:{c['rev']}}},")
    lines.append('];')
    lines.append('')

    # Monthly trend
    months_str = ','.join(f'"{m}"' for m in rev_months)
    labels_str = ','.join(f"'{l}'" for l in rev_labels)
    lines.append(f'const revMonths = [{months_str}];')
    lines.append(f'const revMonthLabels = [{labels_str}];')

    lines.append('const revTrend = {')
    for wine in WINE_ORDER:
        vals = rev_trend.get(wine, [])
        vals_str = ','.join(str(v) for v in vals)
        lines.append(f"  '{wine}': [{vals_str}],")
    lines.append('};')

    lines.append("const revTotalByMonth = revMonths.map((_,i) =>")
    lines.append("  (revTrend['Still Red'][i]||0)+(revTrend['Still White'][i]||0)+(revTrend['Still Rose'][i]||0)+")
    lines.append("  (revTrend['Sparkling White'][i]||0)+(revTrend['Sparkling Rose'][i]||0)")
    lines.append(");")
    lines.append('')
    lines.append(f'const revTotal = {rev_total};')
    lines.append(f'const revUnits = {rev_units};')
    lines.append(f'const revTxns = {rev_txns};')

    # ===== 2026-ONLY REVENUE DATA =====
    lines.append('')
    lines.append('// ==================== 2026-ONLY REVENUE DATA ====================')
    lines.append('const productMix2026 = [')
    if product_mix_2026:
        for p in product_mix_2026:
            lines.append(f"  {{name:'{p['name']}',rev:{p['rev']},units:{p['units']},avgCase:{p['avgCase']}}},")
    lines.append('];')
    lines.append(f'const revTotal2026 = {rev_total_2026};')
    lines.append(f'const revUnits2026 = {rev_units_2026};')
    lines.append(f'const revTxns2026 = {rev_txns_2026};')

    # 2026 monthly wine trend
    if rev_months_2026 and rev_trend_2026:
        m_str = ','.join(f"'{m}'" for m in rev_months_2026)
        l_str = ','.join(f"'{l}'" for l in rev_labels_2026)
        lines.append(f'const revMonths2026 = [{m_str}];')
        lines.append(f'const revMonthLabels2026 = [{l_str}];')
        lines.append('const revTrend2026 = {')
        for wine in WINE_ORDER:
            vals = rev_trend_2026.get(wine, [])
            vals_str = ','.join(str(v) for v in vals)
            lines.append(f"  '{wine}': [{vals_str}],")
        lines.append('};')
    else:
        lines.append('const revMonths2026 = [];')
        lines.append('const revMonthLabels2026 = [];')
        lines.append('const revTrend2026 = {};')

    # ===== 2026 YTD CHANNEL REVENUE =====
    lines.append('')
    lines.append('// ==================== 2026 YTD REVENUE BY CHANNEL ====================')
    channel_order = ['Distributors', 'Website / DTC', 'Direct to Trade - Off Premise', 'Direct to Trade - On Premise', 'Promo', 'Operations', 'Other']
    if channel_ytd:
        all_ytd_months = sorted(set(m for ch in channel_ytd.values() for m in ch))
    else:
        all_ytd_months = []
    ytd_months_str = ','.join(f"'{m}'" for m in all_ytd_months)
    lines.append(f'const ytdMonths = [{ytd_months_str}];')
    # Month labels like "Jan", "Feb"
    ytd_labels = []
    for m in all_ytd_months:
        dt = datetime.strptime(m + '-01', '%Y-%m-%d')
        ytd_labels.append(dt.strftime('%b'))
    ytd_labels_str = ','.join(f"'{l}'" for l in ytd_labels)
    lines.append(f'const ytdMonthLabels = [{ytd_labels_str}];')
    lines.append('const ytdChannelRev = {')
    for ch in channel_order:
        if channel_ytd and ch in channel_ytd:
            vals = [channel_ytd[ch].get(m, 0) for m in all_ytd_months]
            vals_str = ','.join(str(v) for v in vals)
            lines.append(f"  '{ch}': [{vals_str}],")
    lines.append('};')

    # Budget (annual, spread evenly by month)
    lines.append('const ytdBudget = {')
    budget_annual = {
        'Distributors': 1800000,
        'Website / DTC': 500000,
        'Direct to Trade - Off Premise': 500000,
        'Direct to Trade - On Premise': 50000,
    }
    for ch, annual in budget_annual.items():
        monthly = round(annual / 12, 2)
        vals_str = ','.join(str(monthly) for _ in all_ytd_months)
        lines.append(f"  '{ch}': [{vals_str}],")
    lines.append('};')
    budget_total = sum(budget_annual.values())
    lines.append(f'const ytdBudgetTotal = {budget_total};')
    lines.append('')

    # ===== YTD DISTRIBUTOR REVENUE BY STATE =====
    lines.append('// ==================== YTD DISTRIBUTOR REVENUE BY STATE ====================')
    lines.append('const ytdDistRevByState = {')
    if channel_ytd_by_state:
        for st in sorted(channel_ytd_by_state.keys()):
            vals = [channel_ytd_by_state[st].get(m, 0) for m in all_ytd_months]
            vals_str = ','.join(str(v) for v in vals)
            lines.append(f"  '{st}': [{vals_str}],")
    lines.append('};')
    lines.append('')

    # ===== BUILD METADATA =====
    lines.append('')
    # ===== QB DISTRIBUTOR ORDERS =====
    lines.append('// ==================== QB DISTRIBUTOR ORDERS ====================')
    lines.append('const qbDistOrders = {')
    if qb_dist_orders:
        for dist_name, info in sorted(qb_dist_orders.items()):
            lines.append(f"  '{js_str(dist_name)}':{{date:'{info['last_date']}',num:'{info['last_num']}',total:{info['last_total']},items:'{js_str(info['last_items'])}',orders:{info['order_count']},label:'{info.get('label_status','Unknown')}',classicCases:{info.get('classic_cases',0)},contemporaryCases:{info.get('contemporary_cases',0)}}},")
    lines.append('};')
    lines.append('')

    # ===== WAREHOUSE INVENTORY =====
    lines.append('// ==================== WAREHOUSE INVENTORY ====================')
    if warehouse_inv:
        lines.append('const warehouseInventory = {')
        for section in ['classic', 'contemporary']:
            lines.append(f"  {section}: {{")
            for loc, vals in warehouse_inv.get(section, {}).items():
                sw = vals.get('Still White', 0)
                sr = vals.get('Still Rose', 0)
                rd = vals.get('Red', 0)
                spw = vals.get('Sparkling White', 0)
                spr = vals.get('Sparkling Rose', 0)
                t = vals.get('total', 0)
                lines.append(f"    '{js_str(loc)}':{{sw:{sw},sr:{sr},rd:{rd},spw:{spw},spr:{spr},total:{t}}},")
            lines.append('  },')
        # Grand total
        gt = warehouse_inv.get('grand_total', {})
        lines.append(f"  grandTotal:{{sw:{gt.get('Still White',0)},sr:{gt.get('Still Rose',0)},rd:{gt.get('Red',0)},spw:{gt.get('Sparkling White',0)},spr:{gt.get('Sparkling Rose',0)},total:{gt.get('total',0)}}},")
        # Batch 4 estimates
        b4 = warehouse_inv.get('batch4_estimates', {})
        lines.append(f"  batch4Est:{{sw:{b4.get('Still White',0)},sr:{b4.get('Still Rose',0)},rd:{b4.get('Red',0)},spw:{b4.get('Sparkling White',0)},spr:{b4.get('Sparkling Rose',0)},total:{b4.get('total',0)}}},")
        lines.append('};')
    else:
        lines.append('const warehouseInventory = null;')
    lines.append('')

    # ===== CLASSIC INVENTORY TRACKER =====
    lines.append('// ==================== CLASSIC INVENTORY SELLOUT TRACKER ====================')
    if classic_tracker:
        ct = classic_tracker
        lines.append('const classicTracker = {')
        lines.append(f"  trackable: {ct['classic_trackable']},")
        lines.append(f"  totalWithBad: {ct['classic_total']},")
        lines.append(f"  badSpr: {ct['bad_spr']},")
        lines.append(f"  burnRate: {ct['burn_rate_monthly']},")
        lines.append(f"  monthsRemaining: {ct['months_remaining']},")
        lines.append(f"  requiredMonthly: {ct['required_monthly']},")
        lines.append(f"  dataDate: '{ct['data_date']}',")
        # By-wine breakdown
        bw = ct['classic_by_wine']
        lines.append(f"  byWine: {{sw:{bw.get('Still White',0)},sr:{bw.get('Still Rose',0)},rd:{bw.get('Red',0)},spw:{bw.get('Sparkling White',0)},spr:{bw.get('Sparkling Rose',0)}}},")
        # Location details
        lines.append('  locations: {')
        for loc, vals in ct['classic_locations'].items():
            sw = round(vals.get('Still White', 0), 1)
            sr = round(vals.get('Still Rose', 0), 1)
            rd = round(vals.get('Red', 0), 1)
            spw = round(vals.get('Sparkling White', 0), 1)
            spr = round(vals.get('Sparkling Rose', 0), 1)
            t = round(vals.get('total', 0), 1)
            lines.append(f"    '{js_str(loc)}':{{sw:{sw},sr:{sr},rd:{rd},spw:{spw},spr:{spr},total:{t}}},")
        lines.append('  },')
        # History for burn-down chart
        lines.append('  history: [')
        for h in ct['history']:
            lines.append(f"    {{date:'{h['date']}',total:{round(h['total'],1)}}},")
        lines.append('  ],')
        # New label totals for context
        lines.append(f"  newLabelTotal: {ct['new_label_total']},")
        lines.append('};')
    else:
        lines.append('const classicTracker = null;')
    lines.append('')

    lines.append('// ==================== BUILD METADATA ====================')
    build_date = datetime.now().strftime('%B %d, %Y')
    # Data-through date from last week header
    last_week = week_headers[-1] if week_headers else ''
    lines.append(f"const buildDate = '{build_date}';")
    lines.append(f"const dataThrough = '{last_week}';")

    return '\n'.join(lines)


# ============================================================
# HTML INJECTION
# ============================================================

def inject_data_into_html(template_path, js_data, output_path):
    """Replace data section in HTML template with fresh data."""
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    # Find data markers
    start_marker = '// __DATA_START__'
    end_marker = '// __DATA_END__'

    start_idx = html.find(start_marker)
    end_idx = html.find(end_marker)

    if start_idx == -1 or end_idx == -1:
        print("ERROR: Data markers not found in HTML template!")
        print("Make sure the template has '// __DATA_START__' and '// __DATA_END__' markers.")
        sys.exit(1)

    # Replace everything between markers (inclusive of marker lines)
    new_html = html[:start_idx] + start_marker + '\n' + js_data + '\n' + end_marker + html[end_idx + len(end_marker):]

    # Update the date in the header
    today_str = datetime.now().strftime('%b %d, %Y')
    new_html = re.sub(
        r'Data as of <strong>[^<]+</strong>',
        f'Data as of <strong>{today_str}</strong>',
        new_html
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(new_html)

    print(f"Dashboard written to: {output_path}")
    return output_path


# ============================================================
# MAIN
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Rebuild Missing Thorn BI Dashboard from Excel exports')
    parser.add_argument('--data', '--base-dir', dest='data', default=None,
                        help='Path to directory containing Excel source files')
    parser.add_argument('--template', default=None, help='Path to HTML template')
    parser.add_argument('--output', default=None, help='Output path for generated HTML')
    args = parser.parse_args()

    # Determine paths relative to repo root (one level up from scripts/)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(script_dir)

    # Data directory: --data flag > data/ in repo root > same dir as script (legacy)
    base_dir = args.data
    if not base_dir:
        repo_data = os.path.join(repo_root, 'data')
        if os.path.isdir(repo_data):
            base_dir = repo_data
        else:
            base_dir = script_dir
    if not os.path.isdir(base_dir):
        print(f"ERROR: Data directory not found: {base_dir}")
        print("Provide --data flag pointing to your Excel files directory.")
        sys.exit(1)

    # Template: --template flag > templates/executive_dashboard.html in repo root
    template = args.template
    if not template:
        repo_template = os.path.join(repo_root, 'templates', 'executive_dashboard.html')
        if os.path.exists(repo_template):
            template = repo_template
        else:
            # Legacy: look for most recent *_MT_BI_Dashboard.html in script dir
            candidates = [f for f in os.listdir(script_dir) if f.endswith('_MT_BI_Dashboard.html')]
            if candidates:
                template = os.path.join(script_dir, sorted(candidates)[-1])

    if not template or not os.path.exists(template):
        print(f"ERROR: No HTML template found. Provide --template path.")
        sys.exit(1)

    print(f"Data directory: {base_dir}")
    print(f"Template: {template}")

    # ---- Load all Excel data ----
    print("Loading 13W velocity data...")
    vel_grand, vel_dist, vel_dist_sku, week_headers = load_13w_velocity(
        os.path.join(base_dir, 'MT_13W_Distributor_SKU_Velocity.xlsx'))

    print("Loading 4W comparison data...")
    four_w_grand, four_w_dist = load_4w_comparison(
        os.path.join(base_dir, 'MT_4W_vs_Prior4W_SKU.xlsx'))

    print("Loading inventory data...")
    inv_dist, inv_dist_sku = load_inventory(
        os.path.join(base_dir, 'MT_Inventory_DaysOnHand.xlsx'))

    print("Loading placement data...")
    new_dist, new_accounts_raw, lost_dist, sample_placements = load_placements(base_dir)

    print("Loading 4M rolling period data...")
    roll_accounts, roll_account_skus, dist_sku_detail, roll_account_sku_detail, sample_accounts = load_rolling_period(
        os.path.join(base_dir, 'MT_4M_RollingPeriod.xlsx'))

    print(f"  Samples excluded: {len(sample_placements)} placement entries, {len(sample_accounts)} account entries")

    # --- 2025 Archive: used for all-time charts, SKU detail, product mix ---
    print("Loading QuickBooks revenue data...")
    qb_2025 = os.path.join(base_dir, 'QB_Sales_2025_Archive.xlsx')
    qb_legacy1 = os.path.join(base_dir, 'YOLO_SalesByCustomerDetail.xlsx')
    qb_legacy2 = os.path.join(base_dir, 'Sales by Customer Detail - RD.xlsx')
    qb_file = None
    for candidate in [qb_2025, qb_legacy1, qb_legacy2]:
        if os.path.exists(candidate):
            try:
                result = load_quickbooks_revenue(candidate)
                product_mix, cust_rev, monthly_rev, rev_total, rev_units, rev_txns = result[0], result[1], result[2], result[3], result[4], result[5]
                qb_file = candidate
                print(f"  Loaded 2025 archive from {os.path.basename(candidate)}")
                break
            except Exception as e:
                print(f"  {os.path.basename(candidate)} failed ({e}), trying next...")
    if not qb_file:
        print("  WARNING: No 2025 QB archive found")
        product_mix, cust_rev, monthly_rev, rev_total, rev_units, rev_txns = [], [], {}, 0, 0, 0
    dist_cust_ytd_2026 = {}

    # --- 2026 Weekly Export: used for YTD channel vs budget ---
    qb_2026_candidates = [
        os.path.join(base_dir, 'Yolo Brand Group, LLC_Sales by Customer Detail - RD.xlsx'),
        os.path.join(base_dir, 'QB_Sales_2026.xlsx'),
    ]
    channel_ytd = {}
    product_mix_2026 = []
    monthly_rev_2026 = {}
    rev_total_2026 = 0
    rev_units_2026 = 0
    rev_txns_2026 = 0
    qb_2026_file = None
    for candidate in qb_2026_candidates:
        if os.path.exists(candidate):
            try:
                result = load_quickbooks_revenue(candidate)
                channel_ytd = result[6]
                product_mix_2026 = result[7]
                monthly_rev_2026 = result[8]
                rev_total_2026 = result[9]
                rev_units_2026 = result[10]
                rev_txns_2026 = result[11]
                dist_cust_ytd_2026 = result[12]
                qb_2026_file = candidate
                print(f"  Loaded 2026 data from {os.path.basename(candidate)}")
                print(f"    2026 YTD Revenue: ${rev_total_2026:,.2f} | SKUs: {len(product_mix_2026)}")
                break
            except Exception as e:
                print(f"  {os.path.basename(candidate)} failed ({e}), trying next...")
    if not qb_2026_file:
        print("  WARNING: No 2026 QB export found, YTD channel data will be empty")

    print("Loading QuickBooks distributor orders...")
    # Use 2026 file for distributor orders if available, else fall back to archive
    qb_orders_file = qb_2026_file or qb_file
    raw_qb_orders = load_qb_distributor_orders(qb_orders_file) if qb_orders_file else {}
    dep_dist_names = list(set(a['dist'] for a in roll_accounts.values()))
    qb_dist_orders = {}
    for qb_name, order_info in raw_qb_orders.items():
        matched = match_qb_dist_name(qb_name, dep_dist_names)
        for dep_name in matched:
            qb_dist_orders[dep_name] = order_info
    print(f"  Matched {len(qb_dist_orders)} distributors to QB orders")

    # Map QB distributor customer revenue to states
    channel_ytd_by_state = {}  # {state: {month: amount}} for distributor revenue
    if dist_cust_ytd_2026:
        matched_cust = 0
        for qb_cust, month_rev in dist_cust_ytd_2026.items():
            matched_deps = match_qb_dist_name(qb_cust, dep_dist_names)
            if matched_deps:
                # Use the first match's state
                st = get_state_from_dist(matched_deps[0])
                if st:
                    matched_cust += 1
                    if st not in channel_ytd_by_state:
                        channel_ytd_by_state[st] = {}
                    for m, amt in month_rev.items():
                        channel_ytd_by_state[st][m] = channel_ytd_by_state[st].get(m, 0) + amt
        # Round values
        for st in channel_ytd_by_state:
            channel_ytd_by_state[st] = {m: round(v, 2) for m, v in sorted(channel_ytd_by_state[st].items())}
        print(f"  Mapped {matched_cust}/{len(dist_cust_ytd_2026)} QB distributor customers to states for revenue filtering")

    print("Loading warehouse inventory...")
    inv_tracker_file_wh = os.path.join(base_dir, 'Inventory Tracker Live.xlsx')
    warehouse_inv = None
    if os.path.exists(inv_tracker_file_wh):
        warehouse_inv = load_warehouse_inventory(inv_tracker_file_wh)
        classic_total = warehouse_inv.get('classic', {}).get('Warehouse Total', {}).get('total', 0)
        contemp_total = warehouse_inv.get('contemporary', {}).get('Warehouse Total', {}).get('total', 0)
        print(f"  Classic stock: {classic_total:,.0f} cases | Contemporary stock: {contemp_total:,.0f} cases")
        print(f"  Data from tab: {warehouse_inv.get('data_date', 'unknown')}")
    else:
        print("  Inventory Tracker Live.xlsx not found, skipping warehouse data")

    print("Loading Inventory Tracker Live...")
    inv_tracker_file = os.path.join(base_dir, 'Inventory Tracker Live.xlsx')
    classic_tracker = None
    if os.path.exists(inv_tracker_file):
        try:
            classic_tracker = load_inventory_tracker_live(inv_tracker_file)
            print(f"  Classic trackable: {classic_tracker['classic_trackable']:,.1f} cases (excl {classic_tracker['bad_spr']:.0f} bad SpR)")
            print(f"  Burn rate: {classic_tracker['burn_rate_monthly']:,.1f} cases/mo | Months left: {classic_tracker['months_remaining']:.1f}")
            print(f"  New Labels: {classic_tracker['new_label_total']:,.1f} cases")
        except Exception as e:
            print(f"  Failed to load Inventory Tracker Live: {e}")
    else:
        print("  Inventory Tracker Live.xlsx not found, skipping Classic tracker")

    # ---- Transform data ----
    print("Building dashboard data...")
    scorecard = build_dist_scorecard(vel_dist, four_w_dist, inv_dist, new_dist, lost_dist)
    scorecard = consolidate_scorecard(scorecard)
    inventory = build_inventory_data(inv_dist, inv_dist_sku)
    inventory = consolidate_inventory(inventory)
    placements = build_placement_data(new_dist, lost_dist)
    placements = consolidate_placements(placements)
    new_accts = build_new_accounts_list(new_accounts_raw)
    accounts_top = build_accounts_top(roll_accounts, roll_account_skus)
    reorder = build_reorder_data(roll_accounts, roll_account_sku_detail)
    dist_detail = build_dist_detail(vel_dist, vel_dist_sku, four_w_dist, dist_sku_detail)
    rev_months, rev_labels, rev_trend, rev_totals = build_rev_trend(monthly_rev)

    # Build 2026-only revenue trend
    rev_months_2026_out, rev_labels_2026_out, rev_trend_2026_out = [], [], {}
    if monthly_rev_2026:
        rev_months_2026_out, rev_labels_2026_out, rev_trend_2026_out, _ = build_rev_trend(monthly_rev_2026)

    # ---- Generate JS data ----
    print("Generating JavaScript data...")
    js_data = generate_js_data(
        scorecard, inventory, placements, new_accts, accounts_top,
        reorder, dist_detail, product_mix, cust_rev,
        rev_months, rev_labels, rev_trend, rev_totals,
        rev_total, rev_units, rev_txns, week_headers, qb_dist_orders, warehouse_inv,
        channel_ytd, product_mix_2026,
        rev_months_2026_out, rev_labels_2026_out, rev_trend_2026_out,
        rev_total_2026, rev_units_2026, rev_txns_2026,
        classic_tracker, sample_placements, sample_accounts,
        channel_ytd_by_state)

    # ---- Inject into HTML ----
    output = args.output
    if not output:
        output = os.path.join(repo_root, 'dist', 'index.html')

    os.makedirs(os.path.dirname(output) or '.', exist_ok=True)
    result = inject_data_into_html(template, js_data, output)

    print(f"\nDashboard rebuild complete!")
    print(f"  Distributors: {len(scorecard)}")
    print(f"  Inventory entries: {len(inventory)}")
    print(f"  Top accounts: {len(accounts_top)}")
    print(f"  Reorder forecast: {len(reorder)}")
    print(f"  Revenue total: ${rev_total:,.2f}")


if __name__ == '__main__':
    main()
