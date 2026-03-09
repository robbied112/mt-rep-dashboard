#!/usr/bin/env python3
"""
Missing Thorn REP Dashboard - Auto-Rebuild Script
Reads the same Excel exports as build_dashboard.py but generates a REP-FACING dashboard
that EXCLUDES revenue data and REFRAMES negative metrics as opportunities.

Usage:
  python3 build_rep_dashboard.py --data ~/OneDrive/MT\ Dashboard\ Data/ --template templates/rep_dashboard_template.html --output dist/index_rep.html
  python3 build_rep_dashboard.py --base-dir /path/to/Raw\ Exports  (legacy flag, same as --data)
"""
import openpyxl
import json
import re
import sys
import os
from datetime import datetime
from collections import defaultdict

# ============================================================
# CONFIGURATION
# ============================================================
DEFAULT_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '')
DEFAULT_TEMPLATE = None  # Will look for most recent *_MT_BI_RepDashboard.html
DEFAULT_OUTPUT = None     # Will create dated file + index.html

# Wine name mapping (VIP long names -> dashboard short names)
VIP_WINE_MAP = {
    'Missing Thorn Alcohol Removed Red Wine 12/750 ml': 'Still Red',
    'Missing Thorn Alcohol Removed Still White Wine 12/750 ml': 'Still White',
    'Missing Thorn Alcohol Removed Still Rose 12/750 ml': 'Still Rose',
    'Missing Thorn Alcohol Removed Sparkling White Wine 12/750 ml': 'Sparkling White',
    'Missing Thorn Alcohol Removed Sparkling Rose Wine 12/750 ml': 'Sparkling Rose',
}
WINE_ORDER = ['Still Red', 'Still White', 'Still Rose', 'Sparkling White', 'Sparkling Rose']

# Distributors whose pricing is NOT reported
NO_PRICE_PATTERNS = [
    'Republic',
]
NO_PRICE_DISTRIBUTORS = [
    'Dallas',
    'Houston',
]

# Short names for placement data
DIST_SHORT_NAMES = {
    'Breakthru Beverage Florida - Miramar': 'Breakthru FL Mir',
    'Breakthru Beverage Florida - Riverview': 'Breakthru FL Riv',
    'Breakthru Beverage South Carolina - Columbia': 'Breakthru SC',
    'Breakthru Beverage Connecticut - North Haven': 'Breakthru CT',
    'Breakthru Beverage Nevada - Las Vegas': 'Breakthru NV LV',
    'Breakthru Beverage Nevada - Reno': 'Breakthru NV Reno',
    'Fedway Associates - Kearny': 'Fedway NJ',
    'Republic National Distributing Company - Dallas/FTW': 'Republic DAL TX',
    'Republic National Distributing Company - Houston': 'Republic HOU TX',
    'Republic National Distributing Company - San Antonio': 'Republic SA TX',
    'Republic National Distributing Company - Atlanta': 'Republic ATL',
    'Tryon Distributing Company - Durham': 'Tryon Durham NC',
    'Tryon Distributing Co': 'Tryon Co NC',
    'Quail Distributing Company': 'Quail AZ',
    'Pine State Trading Co': 'Pine State ME',
    'Shangy Inc.': 'Shangy PA',
}

# Multi-warehouse distributor groups -- these share inventory within a state
MULTI_WAREHOUSE_GROUPS = {
    'Breakthru FL (Combined)': {
        'st': 'FL',
        'members': ['Breakthru FL - Miramar', 'Breakthru FL - Riverview'],
    },
    'RNDC TX (Combined)': {
        'st': 'TX',
        'members': ['Republic National Dist - Dallas', 'Republic National Dist - Houston',
                     'Republic National Dist - San Antonio'],
    },
    'Winebow CA (Combined)': {
        'st': 'CA',
        'members': ['Winebow - CA North', 'Winebow - CA South'],
    },
}

_DIST_TO_GROUP = {}
for _grp_name, _grp in MULTI_WAREHOUSE_GROUPS.items():
    for _member in _grp['members']:
        _DIST_TO_GROUP[_member] = _grp_name

def get_warehouse_group(dist_name):
    """Return parent group name if dist is in a multi-warehouse group, else None."""
    for prefix, grp_name in _DIST_TO_GROUP.items():
        if dist_name.startswith(prefix):
            return grp_name
    return None

# State name lookup
STATE_NAMES = {
    'NJ': 'New Jersey', 'FL': 'Florida', 'PA': 'Pennsylvania',
    'TX': 'Texas', 'NC': 'North Carolina', 'AZ': 'Arizona',
    'GA': 'Georgia', 'NV': 'Nevada', 'CT': 'Connecticut',
    'ME': 'Maine', 'SC': 'South Carolina', 'CA': 'California',
}

# Region mapping for East/West filter
REGION_MAP = {
    'NJ': 'East', 'FL': 'East', 'PA': 'East', 'NC': 'East',
    'GA': 'East', 'CT': 'East', 'ME': 'East', 'SC': 'East',
    'TX': 'West', 'AZ': 'West', 'NV': 'West', 'CA': 'West',
}


# ============================================================
# UTILITY FUNCTIONS
# ============================================================
def clean_dist(name):
    if not name or name == 'Total':
        return name
    if 'Report Created' in str(name):
        return None
    return str(name).strip()

def map_wine(vip_name):
    return VIP_WINE_MAP.get(str(vip_name), str(vip_name))

def safe_num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def safe_pct(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return None

def is_sample_account(name):
    """Detect sample/marketing accounts in VIP data.
    Rule: account name contains 'SAMPLE' (case-insensitive)."""
    return 'sample' in str(name).lower()

def is_likely_personal_sample(name, revenue):
    """Detect likely personal/sample entries in QB data.
    Rule: if name looks like a person (2-3 words, all alpha, no business keywords)
    and revenue is $0 or negligible, it's likely a sample."""
    if revenue > 10:
        return False
    name = str(name).strip()
    words = name.split()
    if len(words) < 2 or len(words) > 4:
        return False
    biz_words = {'inc', 'llc', 'corp', 'company', 'co', 'distribut', 'beverage',
                 'wine', 'spirits', 'store', 'market', 'restaurant', 'bar', 'hotel',
                 'group', 'trading', 'associates', 'wholesale', 'cellar', 'shop',
                 'cafe', 'bistro', 'grill', 'pub', 'lounge', 'thorn', 'yolo'}
    name_lower = name.lower()
    for bw in biz_words:
        if bw in name_lower:
            return False
    alpha_words = sum(1 for w in words if w.replace('.','').replace(',','').isalpha())
    return alpha_words >= 2

def js_str(s):
    """Escape a string for JS single-quoted literal."""
    return str(s).replace("\\", "\\\\").replace("'", "\\'")

def parse_date_str(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), '%m/%d/%Y')
    except:
        return None

def get_dist_short(dist_name):
    """Get short distributor name for placement charts."""
    for long, short in DIST_SHORT_NAMES.items():
        if long.lower() in dist_name.lower() or dist_name.lower() in long.lower():
            return short
    return dist_name[:20] if len(dist_name) > 20 else dist_name

def get_state_from_dist(dist_name):
    """Extract state abbreviation from distributor name suffix."""
    parts = dist_name.split(',')
    if len(parts) >= 2:
        st = parts[-1].strip().upper()
        if len(st) == 2:
            return st
    for st, name in STATE_NAMES.items():
        if f', {st}' in dist_name or f'- {name}' in dist_name:
            return st
    return ''


# ============================================================
# MOMENTUM CALCULATION (REP-FOCUSED REFRAMING)
# ============================================================

def fmt_momentum(prior, current):
    """
    Format momentum for rep dashboard.
    - If current > prior: show positive percentage
    - If current < prior: show "Opportunity: +X CE" (the gap)
    - If prior is 0: show "New Market"
    """
    if prior is None or prior < 2.0:
        # Treat prior < 2.0 CE as effectively "New Market" -- tiny prior values
        # (e.g. 0.3 CE) produce misleading 1000%+ growth percentages
        return 'New Market'

    if current >= prior:
        # Positive or flat
        pct = ((current - prior) / abs(prior)) * 100
        return f'+{pct:.1f}%' if pct >= 0 else f'{pct:.1f}%'
    else:
        # Negative - reframe as opportunity
        gap = prior - current
        return f'Opportunity: +{gap:.0f} CE'


def compute_consistency_label(con):
    """
    Reframe consistency as account maturity stage.
    - < 0.5 = "Emerging" (bought but not regular)
    - 0.5-0.8 = "Building"
    - >= 0.8 = "Established"
    """
    if con < 0.5:
        return 'Emerging'
    elif con < 0.8:
        return 'Building'
    else:
        return 'Established'


def compute_trend_label(first_half, second_half):
    """
    Reframe account trends for reps.
    - If accelerating: "Momentum"
    - If decelerating: "Growth Opportunity"
    - If steady: "Consistent"
    """
    if first_half == 0 and second_half > 0:
        return 'Momentum'
    elif second_half > first_half * 1.1:
        return 'Momentum'
    elif second_half < first_half * 0.7:
        return 'Growth Opportunity'
    else:
        return 'Consistent'


def compute_inventory_status_rep(doh):
    """
    Reframe inventory status for reps (opportunities vs warnings).
    - <= 0 = "Review Needed"
    - 1-90 = "Reorder Opportunity"
    - 91-180 = "Healthy"
    - 181-365 = "Overstocked"
    - 365+ = "Dead Stock"
    """
    if doh <= 0:
        return 'Review Needed'
    if doh <= 90:
        return 'Reorder Opportunity'
    if doh <= 180:
        return 'Healthy'
    if doh <= 365:
        return 'Overstocked'
    return 'Dead Stock'  # 365+ days


# ============================================================
# EXCEL PARSING (SAME AS build_dashboard.py)
# ============================================================

def load_13w_velocity(filepath):
    """Parse MT_13W_DistributorSKUVelocity.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    week_cols_ce = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28]

    # Grand total from row 3
    grand_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    grand = {
        '13w_ce': safe_num(grand_row[30]),
        'prior_ce': safe_num(grand_row[32]),
        'weekly_ce': [safe_num(grand_row[c]) for c in week_cols_ce],
    }

    # Week headers from row 1
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    week_headers = []
    for c in week_cols_ce:
        h = row1[c]
        if h:
            week_headers.append(str(h).replace('1 Week ', ''))
        else:
            week_headers.append('')

    vel_dist = {}
    vel_dist_sku = {}

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        item = str(row[1]) if row[1] else ''
        state = str(row[2]) if row[2] else ''
        ym = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if item == 'Total' and state == 'Total' and ym == 'Total':
            weekly = [safe_num(row[c]) for c in week_cols_ce]
            vel_dist[dist] = {
                '13w_ce': safe_num(row[30]),
                'prior_ce': safe_num(row[32]),
                'diff_ce': safe_num(row[34]),
                'pct_ce': safe_pct(row[36]),
                'weekly_ce': weekly,
            }
        elif item != 'Total' and state == 'Total' and ym == 'Total':
            wine = map_wine(item)
            weekly = [safe_num(row[c]) for c in week_cols_ce]
            vel_dist_sku[(dist, wine)] = {
                '13w_ce': safe_num(row[30]),
                'prior_ce': safe_num(row[32]),
                'weekly_ce': weekly,
            }

    wb.close()
    return grand, vel_dist, vel_dist_sku, week_headers


def load_4w_comparison(filepath):
    """Parse MT_4Wvs4W_SKU.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    grand_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    four_w_grand = {
        'cur_ce': safe_num(grand_row[3]),
        'prior_ce': safe_num(grand_row[5]),
    }

    four_w_dist = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        item = str(row[2]) if row[2] else ''

        if not dist or dist == 'Total':
            continue
        if state == 'Total' and item == 'Total':
            four_w_dist[dist] = {
                'cur_ce': safe_num(row[3]),
                'prior_ce': safe_num(row[5]),
            }

    wb.close()
    return four_w_grand, four_w_dist


def load_inventory(filepath):
    """Parse MT_Inventory_DaysOnHand.xlsx"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    inv_dist = {}
    inv_dist_sku = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        item = str(row[2]) if row[2] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and item == 'Total':
            inv_dist[dist] = {
                'on_hand': safe_num(row[3]),
                'daily_rate': safe_num(row[4]),
                'doh': safe_num(row[5]),
                '90d_ce': safe_num(row[6]),
                'proj_order': safe_num(row[11]),
            }
        elif item != 'Total' and state != 'Total':
            wine = map_wine(item)
            inv_dist_sku[(dist, state, wine)] = {
                'on_hand': safe_num(row[3]),
                'daily_rate': safe_num(row[4]),
                'doh': safe_num(row[5]),
            }

    wb.close()
    return inv_dist, inv_dist_sku


def load_placements(base_dir):
    """Parse MT_NewPlacements.xlsx and MT_LostPlacements.xlsx"""
    # New placements
    wb_new = openpyxl.load_workbook(os.path.join(base_dir, 'MT_NewPlacements_30Sold_90Unsold_CaseEq.xlsx'), data_only=True)
    ws_new = wb_new.active

    new_dist = {}
    new_accounts = []
    sample_placements = []

    for row in ws_new.iter_rows(min_row=3, max_row=ws_new.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        acct = str(row[2]) if row[2] else ''
        item = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and acct == 'Total' and item == 'Total':
            new_dist[dist] = {'sold_ce': safe_num(row[4]), 'new_accts': int(safe_num(row[6]))}
        elif acct != 'Total' and item == 'Total':
            placement_entry = {
                'dist': dist, 'state': state, 'account': acct,
                'sold_ce': safe_num(row[4]), 'new_accts': int(safe_num(row[6])),
            }
            if is_sample_account(acct):
                sample_placements.append(placement_entry)
            else:
                new_accounts.append(placement_entry)
    wb_new.close()

    # Lost placements
    wb_lost = openpyxl.load_workbook(os.path.join(base_dir, 'MT_LostPlacements_30Sold_90Unsold.xlsx'), data_only=True)
    ws_lost = wb_lost.active

    lost_dist = {}
    lost_accounts = []  # STORE LOST ACCOUNT DETAILS FOR REP REFRAMING
    for row in ws_lost.iter_rows(min_row=4, max_row=ws_lost.max_row, values_only=True):
        dist = clean_dist(row[0])
        state = str(row[1]) if row[1] else ''
        acct = str(row[2]) if row[2] else ''
        item = str(row[3]) if row[3] else ''

        if not dist or dist == 'Total':
            continue

        if state == 'Total' and acct == 'Total' and item == 'Total':
            lost_dist[dist] = {'prior_ce': safe_num(row[5]), 'lost_accts': int(safe_num(row[7]))}
        elif acct != 'Total' and item == 'Total':
            # Store as re-engagement opportunity
            lost_accounts.append({
                'dist': dist, 'state': state, 'account': acct,
                'prior_ce': safe_num(row[5]), 'lost_accts': int(safe_num(row[7])),
            })
    wb_lost.close()

    return new_dist, new_accounts, lost_dist, lost_accounts, sample_placements


def load_rolling_period(filepath):
    """Parse MT_4M_RollingPeriod.xlsx for account-level data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    roll_accounts = {}
    roll_account_skus = {}
    dist_sku_detail = {}
    roll_account_sku_detail = {}  # NEW: (acct, dist, wine) -> SKU-level data per account
    sample_accounts = {}

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        item = str(row[0]) if row[0] else ''
        dist = str(row[1]).strip() if row[1] else ''
        oo = str(row[2]).strip() if row[2] else ''
        acct = str(row[3]).strip() if row[3] else ''
        state = str(row[4]).strip() if row[4] else ''

        if not acct or acct == 'Total' or not dist or dist == 'Total':
            continue
        if 'Report Created' in item:
            continue
        if oo == 'Total':
            continue
        if 'YOLO' in acct.upper():
            continue

        wine = map_wine(item)

        # Handle sample accounts separately
        if is_sample_account(acct):
            key = (acct, dist)
            if key not in sample_accounts:
                sample_accounts[key] = {
                    'account': acct, 'dist': dist, 'state': state, 'channel': oo,
                    'total_ce': 0, 'total_rev': 0,
                }
            sample_accounts[key]['total_ce'] += safe_num(row[37])
            continue

        # --- Account-level aggregation ---
        key = (acct, dist)
        if key not in roll_accounts:
            roll_accounts[key] = {
                'account': acct, 'dist': dist, 'state': state, 'channel': oo,
                'nov_ce': 0, 'dec_ce': 0, 'jan_ce': 0, 'feb_ce': 0,
                'total_ce': 0, 'total_purchases': 0,
                'prior_ce': 0,
                'first_buy': None, 'last_buy': None,
                'eff_placements': 0,
            }
            roll_account_skus[key] = set()

        a = roll_accounts[key]
        roll_account_skus[key].add(wine)

        a['nov_ce'] += safe_num(row[5])
        a['dec_ce'] += safe_num(row[13])
        a['jan_ce'] += safe_num(row[21])
        a['feb_ce'] += safe_num(row[29])
        a['total_ce'] += safe_num(row[37])
        a['total_purchases'] += safe_num(row[42])
        a['eff_placements'] += safe_num(row[39])
        a['prior_ce'] += safe_num(row[45])

        for col in [11, 12, 19, 20, 27, 28, 35, 36, 43, 44]:
            dt = parse_date_str(row[col])
            if dt:
                if not a['first_buy'] or dt < a['first_buy']:
                    a['first_buy'] = dt
                if not a['last_buy'] or dt > a['last_buy']:
                    a['last_buy'] = dt

        # --- Account+SKU level aggregation (for reorder by SKU) ---
        ask = (acct, dist, wine)
        if ask not in roll_account_sku_detail:
            roll_account_sku_detail[ask] = {
                'ce': 0, 'total_purchases': 0,
                'first_buy': None, 'last_buy': None,
                'nov_ce': 0, 'dec_ce': 0, 'jan_ce': 0, 'feb_ce': 0,
            }
        asd = roll_account_sku_detail[ask]
        asd['ce'] += safe_num(row[37])
        asd['total_purchases'] += safe_num(row[42])
        asd['nov_ce'] += safe_num(row[5])
        asd['dec_ce'] += safe_num(row[13])
        asd['jan_ce'] += safe_num(row[21])
        asd['feb_ce'] += safe_num(row[29])

        for col in [11, 12, 19, 20, 27, 28, 35, 36, 43, 44]:
            dt = parse_date_str(row[col])
            if dt:
                if not asd['first_buy'] or dt < asd['first_buy']:
                    asd['first_buy'] = dt
                if not asd['last_buy'] or dt > asd['last_buy']:
                    asd['last_buy'] = dt

        # --- Distributor+SKU aggregation ---
        dk = (dist, wine)
        if dk not in dist_sku_detail:
            dist_sku_detail[dk] = {
                'ce': 0, 'prior_ce': 0,
            }
        ds = dist_sku_detail[dk]
        ds['ce'] += safe_num(row[37])
        ds['prior_ce'] += safe_num(row[45])

    # Add SKU count
    for key in roll_accounts:
        roll_accounts[key]['sku_count'] = len(roll_account_skus[key])

    wb.close()
    return roll_accounts, roll_account_skus, dist_sku_detail, roll_account_sku_detail, sample_accounts


# ============================================================
# DATA TRANSFORMATION (REP-FOCUSED)
# ============================================================

def build_dist_scorecard_rep(vel_dist, four_w_dist, inv_dist, new_dist, lost_dist):
    """
    Build the distScorecard for rep dashboard.
    - Excludes revenue
    - Uses momentum reframing
    - Includes consistency label
    """
    scorecard = []
    for dist, v in vel_dist.items():
        st = get_state_from_dist(dist)

        # 4W data
        fw = four_w_dist.get(dist, {})
        w4_ce = fw.get('cur_ce', 0)
        prior_w4 = fw.get('prior_ce', 0)

        # Inventory
        inv = inv_dist.get(dist, {})
        oh = inv.get('on_hand', 0)
        doh = inv.get('doh', 0)

        # Placements
        nd = new_dist.get(dist, {})
        ld = lost_dist.get(dist, {})
        net = nd.get('new_accts', 0) - ld.get('lost_accts', 0)

        # Consistency
        weekly = v.get('weekly_ce', [0]*13)
        active_weeks = sum(1 for w in weekly if w > 0)
        con = round(active_weeks / 13, 2)
        con_label = compute_consistency_label(con)

        # Momentum (reframed)
        momentum = fmt_momentum(v.get('prior_ce', 0), v.get('13w_ce', 0))
        w4_momentum = fmt_momentum(prior_w4, w4_ce)

        # Inventory action classification
        if doh <= 0:
            inv_action = 'No Data'
        elif doh <= 45:
            inv_action = 'Reorder Now'
        elif doh <= 90:
            inv_action = 'Monitor'
        elif doh <= 300:
            inv_action = 'Healthy'
        else:
            inv_action = 'Reduce'

        # Velocity trend: compare last 4 weeks vs first 4 weeks of 13W
        first4 = sum(weekly[:4])
        last4 = sum(weekly[-4:])
        if first4 < 0.5:
            vel_trend = 'new'
        elif last4 >= first4 * 1.1:
            vel_trend = 'up'
        elif last4 <= first4 * 0.9:
            vel_trend = 'down'
        else:
            vel_trend = 'flat'

        # Sell-through rate: 13W CE / (13W CE + current OH) * 100
        total_product = v.get('13w_ce', 0) + oh
        sell_through = round((v.get('13w_ce', 0) / total_product) * 100, 1) if total_product > 0 else 0

        scorecard.append({
            'name': dist, 'st': st,
            'ce': round(v.get('13w_ce', 0), 1),
            'prior': round(v.get('prior_ce', 0), 1),
            'momentum': momentum,
            'w4': round(w4_ce, 1),
            'prior_w4': round(prior_w4, 1),
            'w4trend': w4_momentum,
            'oh': round(oh, 1),
            'doh': round(doh),
            'net': net,
            'con': con,
            'conLabel': con_label,
            'weekly': [round(w, 1) for w in weekly],
            'invAction': inv_action,
            'velTrend': vel_trend,
            'sellThru': sell_through,
        })

    scorecard.sort(key=lambda x: -x['ce'])
    return scorecard


def match_qb_dist_name(qb_name, dep_dist_names):
    """Fuzzy match a QuickBooks distributor name to a depletion report distributor name."""
    qb_lower = qb_name.lower().strip()
    keyword_map = [
        ('shangy', 'shangy'),
        ('fedway', 'fedway'),
        ('pine state', 'pine state'),
        ('quail', 'quail'),
        ('tryon', 'tryon'),
        ('okoboji', 'okoboji'),
        ('libdib', 'libdib'),
        ('winebow', 'winebow'),
        ('craft brewers', 'craft brewers'),
        ('ben arnold', 'ben arnold'),
        ('breakthru beverage nevada', 'breakthru nv'),
        ('premier beverage company (tampa)', 'breakthru fl - riverview'),
        ('premiere beverage company (miramar):connecticut', 'breakthru ct'),
        ('premiere beverage company (miramar)', 'breakthru fl - miramar'),
        ('carisam', 'breakthru fl'),
        ('rndc ga', 'republic national dist - atlanta'),
        ('republic national distributing', 'republic national dist'),
        ('aoc', 'aoc'),
    ]
    for qb_kw, dep_kw in keyword_map:
        if qb_kw in qb_lower:
            matches = [d for d in dep_dist_names if dep_kw in d.lower()]
            if matches:
                return matches
    return []


def classify_qb_product_label(product_name):
    """Classify a QB product name as Classic or Contemporary label."""
    p_lower = product_name.lower()
    if any(skip in p_lower for skip in ['shipping', 'crv', 'discount', 'tax', 'sample', 'bottle - 750', 'bottle of', 'bottle missing']):
        return None
    if '12 pack' in p_lower or '12pk' in p_lower:
        return 'Contemporary'
    if 'cases of na wine' in p_lower:
        return 'Classic'
    return None


def load_qb_distributor_orders(filepath):
    """Parse QuickBooks Sales by Customer Detail for distributor order history with label classification."""
    from datetime import datetime
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    invoices = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        ctype = str(row[12]).strip() if row[12] else ''
        if ctype != 'Distributors':
            continue
        date_val = row[1]
        num = str(row[3]).strip() if row[3] else ''
        product = str(row[4]).strip() if row[4] else ''
        qty = row[6] if row[6] else 0
        amount = row[8] if row[8] else 0
        cname = str(row[10]).strip() if row[10] else ''
        if not date_val or not cname or cname == 'None' or not num:
            continue
        if isinstance(date_val, datetime):
            dt = date_val
        else:
            try:
                dt = datetime.strptime(str(date_val), '%m/%d/%Y')
            except:
                continue
        key = (cname, num)
        if key not in invoices:
            invoices[key] = {'date': dt, 'items': [], 'total': 0, 'num': num, 'customer': cname}
        invoices[key]['items'].append({'product': product, 'qty': qty})
        invoices[key]['total'] += amount if amount else 0

    dist_orders = {}
    dist_labels = {}
    for (cname, num), inv in invoices.items():
        if cname not in dist_orders:
            dist_orders[cname] = []
        if cname not in dist_labels:
            dist_labels[cname] = {'Classic': 0, 'Contemporary': 0}

        simple_items = []
        for item in inv['items']:
            p = item['product']
            label_type = classify_qb_product_label(p)
            if label_type and item['qty'] and item['qty'] > 0:
                dist_labels[cname][label_type] += int(item['qty'])

            p_lower = p.lower()
            wine = 'Other'
            if 'red' in p_lower:
                wine = 'Red'
            elif 'sparkling white' in p_lower or 'sparkling whi' in p_lower:
                wine = 'Spark White'
            elif 'sparkling ros' in p_lower:
                wine = 'Spark Rose'
            elif 'still white' in p_lower or ('white' in p_lower and 'sparkling' not in p_lower):
                wine = 'White'
            elif 'still ros' in p_lower or ('ros' in p_lower and 'sparkling' not in p_lower):
                wine = 'Rose'
            if item['qty'] and item['qty'] > 0:
                simple_items.append({'wine': wine, 'qty': int(item['qty'])})
        dist_orders[cname].append({'date': inv['date'], 'num': inv['num'], 'items': simple_items, 'total': inv['total']})

    result = {}
    for cname, orders in dist_orders.items():
        orders.sort(key=lambda x: x['date'], reverse=True)
        latest = orders[0]
        item_summary = {}
        for it in latest['items']:
            w = it['wine']
            if w not in item_summary:
                item_summary[w] = 0
            item_summary[w] += it['qty']
        summary_str = ', '.join([f"{w} x{q}" for w, q in item_summary.items() if w != 'Other'])

        labels = dist_labels.get(cname, {'Classic': 0, 'Contemporary': 0})
        has_classic = labels['Classic'] > 0
        has_contemporary = labels['Contemporary'] > 0
        if has_classic and has_contemporary:
            label_status = 'Both'
        elif has_classic:
            label_status = 'Classic'
        elif has_contemporary:
            label_status = 'Contemporary'
        else:
            label_status = 'Unknown'

        result[cname] = {
            'last_date': latest['date'].strftime('%m/%d/%y'),
            'last_num': latest['num'],
            'last_total': round(latest['total']),
            'last_items': summary_str,
            'order_count': len(orders),
            'label_status': label_status,
            'classic_cases': labels['Classic'],
            'contemporary_cases': labels['Contemporary']
        }

    wb.close()
    return result


def load_inventory_tracker_live(filepath):
    """Parse Inventory Tracker Live.xlsx for Classic (Batch 1+2) sellout tracker.
    Reads the most recent date-stamped sheet, extracts Batch 1+2 and Batch 3 data.
    Excludes 627 bad Sparkling Rose at Zephyr Batch 1.
    Also reads historical sheets to compute burn rate.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find most recent date sheet by parsing M.D.YY dates and sorting
    date_sheets = []
    for s in wb.sheetnames:
        if not any(c.isdigit() for c in s) or '.' not in s:
            continue
        if any(kw in s for kw in ('Summary', 'Sales', 'Change', 'Amazon', 'Cases')):
            continue
        parts = s.split('.')
        if len(parts) >= 3:
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                yr = y + 2000 if y < 100 else y
                date_sheets.append((yr, m, d, s))
            except ValueError:
                pass
    if date_sheets:
        date_sheets.sort(reverse=True)
        latest_sheet = date_sheets[0][3]
    else:
        latest_sheet = '2.9.26'

    ws = wb[latest_sheet]
    wine_names = ['Still White', 'Still Rose', 'Red', 'Sparkling White', 'Sparkling Rose']

    def sv(v):
        if v is None: return 0
        try: return float(v)
        except: return 0

    # Batch 1+2 (Classic) - Rows 4-9, cols F(6)-L(12)
    classic_locations = {}
    for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {
            'Still White': sv(row[6]), 'Still Rose': sv(row[7]),
            'Red': sv(row[8]), 'Sparkling White': sv(row[9]),
            'Sparkling Rose': sv(row[10]), 'total': sv(row[11])
        }
        classic_locations[loc] = vals

    # Batch 3 (New Labels) - Rows 15-19, cols F-L
    new_label_locations = {}
    for row in ws.iter_rows(min_row=15, max_row=19, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {
            'Still White': sv(row[6]), 'Still Rose': sv(row[7]),
            'Red': sv(row[8]), 'Sparkling White': sv(row[9]),
            'Sparkling Rose': sv(row[10]), 'total': sv(row[11])
        }
        new_label_locations[loc] = vals

    # BAD inventory: 627 Sparkling Rose at Zephyr Express, CA Batch 1
    bad_spr = 0
    for loc, vals in classic_locations.items():
        if 'Batch 1' in loc and 'Zephyr' in loc:
            bad_spr = vals.get('Sparkling Rose', 0)

    # Classic totals (from Warehouse Total row)
    classic_total_row = classic_locations.get('Warehouse Total', {})
    classic_total = classic_total_row.get('total', 0)
    classic_trackable = classic_total - bad_spr

    # Classic by wine (trackable - excluding bad SpR)
    classic_by_wine = {}
    for wine in wine_names:
        classic_by_wine[wine] = classic_total_row.get(wine, 0)
    classic_by_wine['Sparkling Rose'] -= bad_spr

    # New label totals
    new_total_row = new_label_locations.get('Warehouse Total', {})
    new_total = new_total_row.get('total', 0)

    # Historical burn rate from previous sheets (use sorted date_sheets)
    history = []
    historical_sheets = [s[3] for s in sorted(date_sheets)]  # chronological order
    for sname in historical_sheets:
        if sname in wb.sheetnames:
            hws = wb[sname]
            for row in hws.iter_rows(min_row=9, max_row=9, values_only=True):
                loc = str(row[5]).strip() if row[5] else ''
                if 'Warehouse Total' in loc:
                    t = sv(row[11])
                    # Parse date from sheet name
                    parts = sname.replace('.25', '.2025').replace('.26', '.2026').replace('.24', '.2024').split('.')
                    if len(parts) >= 3:
                        date_str = f"{parts[0]}/{parts[1]}/{parts[2]}"
                    elif len(parts) == 2:
                        date_str = sname
                    else:
                        date_str = sname
                    history.append({'date': sname, 'total': t})

    # Compute monthly burn rate from last 3 data points
    burn_rate_monthly = 0
    if len(history) >= 2:
        # Use last 3 months of data for burn rate
        recent = history[-3:] if len(history) >= 3 else history
        total_drop = recent[0]['total'] - recent[-1]['total']
        # Estimate months between data points (roughly 1 month apart)
        n_periods = len(recent) - 1
        if n_periods > 0 and total_drop > 0:
            burn_rate_monthly = total_drop / n_periods

    # Deadline: June 30, 2026
    from datetime import datetime
    today = datetime.now()
    deadline = datetime(2026, 6, 30)
    months_remaining = max(0, (deadline - today).days / 30.44)
    required_monthly = classic_trackable / months_remaining if months_remaining > 0 else 0

    wb.close()

    return {
        'classic_locations': {k: v for k, v in classic_locations.items() if k not in ('Warehouse Total', 'Zephyr Total')},
        'classic_total': classic_total,
        'classic_trackable': round(classic_trackable, 1),
        'classic_by_wine': {k: round(v, 1) for k, v in classic_by_wine.items()},
        'bad_spr': round(bad_spr, 1),
        'new_label_locations': {k: v for k, v in new_label_locations.items() if k != 'Warehouse Total'},
        'new_label_total': round(new_total, 1),
        'burn_rate_monthly': round(burn_rate_monthly, 1),
        'months_remaining': round(months_remaining, 1),
        'required_monthly': round(required_monthly, 1),
        'history': history,
        'data_date': latest_sheet,
    }


def load_warehouse_inventory(filepath):
    """Parse Inventory Tracker Live.xlsx for warehouse stock levels.
    Uses the most recent date-stamped tab instead of a fixed summary sheet.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    wine_cols = ['Still White', 'Still Rose', 'Red', 'Sparkling White', 'Sparkling Rose']

    # Find most recent date-stamped sheet (same logic as load_inventory_tracker_live)
    date_sheets = []
    for s in wb.sheetnames:
        if not any(c.isdigit() for c in s) or '.' not in s:
            continue
        if any(kw in s for kw in ('Summary', 'Sales', 'Change', 'Amazon', 'Cases')):
            continue
        parts = s.split('.')
        if len(parts) >= 3:
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                yr = y + 2000 if y < 100 else y
                date_sheets.append((yr, m, d, s))
            except ValueError:
                pass
    if date_sheets:
        date_sheets.sort(reverse=True)
        latest_sheet = date_sheets[0][3]
    else:
        # Fallback to Jan 2026 Summary if no dated sheets
        latest_sheet = 'Jan 2026 Summary'

    print(f"  Using warehouse tab: {latest_sheet}")
    ws = wb[latest_sheet]

    def safe_val(v):
        if v is None or str(v).strip() == 'None':
            return 0
        try:
            return round(float(v))
        except:
            return 0

    # Batch 1+2 (Classic) - Rows 4-9, cols F(6)-L(12)
    classic = {}
    for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
        vals['total'] = safe_val(row[11])
        classic[loc] = vals

    # Batch 3 (Contemporary/New Labels) - Rows 15-20, cols F-L
    contemporary = {}
    for row in ws.iter_rows(min_row=15, max_row=20, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if not loc or loc == 'None':
            continue
        vals = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
        vals['total'] = safe_val(row[11])
        contemporary[loc] = vals

    # Quantity Produced (row 21)
    produced = {}
    for row in ws.iter_rows(min_row=21, max_row=21, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if loc and 'Produced' in loc:
            produced = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            produced['total'] = safe_val(row[11])

    # Grand total (row 24) and Batch 4 estimates (row 25) - scan rows 23-26
    grand_total = {}
    batch4 = {}
    for row in ws.iter_rows(min_row=23, max_row=27, values_only=True):
        loc = str(row[5]).strip() if row[5] else ''
        if 'Total' in loc and 'Batch' not in loc and 'Post' not in loc:
            grand_total = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            grand_total['total'] = safe_val(row[11])
        elif 'Batch 4' in loc:
            batch4 = {wine_cols[i]: safe_val(row[6+i]) for i in range(5)}
            batch4['total'] = safe_val(row[11])

    wb.close()
    return {
        'classic': classic,
        'contemporary': contemporary,
        'batch3_produced': produced,
        'batch4_estimates': batch4,
        'grand_total': grand_total,
        'data_date': latest_sheet,
    }


def build_inventory_data_rep(inv_dist, inv_dist_sku=None):
    """Build inventory data with opportunity-focused status labels, including SKU breakout."""
    inv_data = []
    for dist, inv in inv_dist.items():
        st = get_state_from_dist(dist)
        doh = inv.get('doh', 0)
        status = compute_inventory_status_rep(doh)

        # Build SKU breakout if available
        skus = []
        if inv_dist_sku:
            for wine in WINE_ORDER:
                # inv_dist_sku key is (dist, state, wine)
                for (d, s, w), idata in inv_dist_sku.items():
                    if d == dist and w == wine:
                        sku_oh = round(idata.get('on_hand', 0), 1)
                        sku_rate = round(idata.get('daily_rate', 0), 2)
                        sku_doh = round(idata.get('doh', 0))
                        if sku_oh > 0 or sku_rate > 0:
                            skus.append({
                                'w': wine,
                                'oh': sku_oh,
                                'rate': sku_rate,
                                'doh': sku_doh,
                                'status': compute_inventory_status_rep(sku_doh),
                            })
                        break  # one state per dist per wine typically

        inv_data.append({
            'name': dist, 'st': st,
            'oh': round(inv.get('on_hand', 0), 1),
            'rate': round(inv.get('daily_rate', 0), 2),
            'doh': round(doh),
            'dep90': round(inv.get('90d_ce', 0), 1),
            'status': status,
            'proj': round(inv.get('proj_order', 0), 1),
            'skus': skus,
        })

    inv_data.sort(key=lambda x: -x['oh'])
    return inv_data


def build_reengagement_data(lost_accounts):
    """
    Build re-engagement opportunities (previously lost placements).
    Frame as "doors that already said yes once".
    """
    reeng = []
    for a in lost_accounts:
        # Re-engagement priority (based on prior volume)
        priority = min(100, round(a['prior_ce'] * 10 + a['lost_accts'] * 15))

        reeng.append({
            'account': a['account'],
            'dist': a['dist'],
            'state': a['state'],
            'priorAccts': int(a['lost_accts']),
            'priorCE': round(a['prior_ce'], 1),
            'priority': priority,
        })

    # Sort by prior CE (biggest opportunities first)
    reeng.sort(key=lambda x: -x['priorCE'])
    return reeng


# ============================================================
# MULTI-WAREHOUSE CONSOLIDATION
# ============================================================

def consolidate_scorecard(scorecard):
    """Consolidate multi-warehouse distributors into parent groups with children."""
    groups = {}
    singles = []
    for entry in scorecard:
        grp = get_warehouse_group(entry['name'])
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)
    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        ce = sum(c['ce'] for c in children)
        rev = sum(c.get('rev', 0) for c in children)
        prior = sum(c['prior'] for c in children)
        w4 = sum(c['w4'] for c in children)
        prior_w4 = sum(c.get('prior_w4', 0) for c in children)
        oh = sum(c['oh'] for c in children)
        net = sum(c['net'] for c in children)
        total_rate = sum(c['oh'] / c['doh'] if c['doh'] > 0 else 0 for c in children)
        doh = oh / total_rate if total_rate > 0 else 0
        con = max(c['con'] for c in children)
        weekly = [0.0] * 13
        for c in children:
            for i, w in enumerate(c.get('weekly', [0]*13)):
                weekly[i] += w
        # Use fmt_momentum for rep dashboard instead of fmt_yoy/fmt_mom
        momentum = fmt_momentum(prior, ce)
        w4_momentum = fmt_momentum(prior_w4, w4)

        # Velocity trend: compare last 4 weeks vs first 4 weeks of 13W
        first4 = sum(weekly[:4])
        last4 = sum(weekly[-4:])
        if first4 < 0.5:
            vel_trend = 'new'
        elif last4 >= first4 * 1.1:
            vel_trend = 'up'
        elif last4 <= first4 * 0.9:
            vel_trend = 'down'
        else:
            vel_trend = 'flat'

        # Sell-through rate: 13W CE / (13W CE + current OH) * 100
        total_product = ce + oh
        sell_through = round((ce / total_product) * 100, 1) if total_product > 0 else 0

        result.append({
            'name': grp_name, 'st': cfg['st'],
            'ce': round(ce, 1), 'prior': round(prior, 1),
            'momentum': momentum,
            'w4': round(w4, 1), 'prior_w4': round(prior_w4, 1),
            'w4trend': w4_momentum,
            'oh': round(oh, 1), 'doh': round(doh),
            'net': net, 'con': round(con, 2),
            'conLabel': compute_consistency_label(con),
            'weekly': [round(w, 1) for w in weekly],
            'velTrend': vel_trend,
            'sellThru': sell_through,
            'isGroup': True,
            'children': sorted(children, key=lambda x: -x['ce']),
        })
    result.extend(singles)
    result.sort(key=lambda x: -x['ce'])
    return result

def consolidate_inventory(inv_data):
    """Consolidate multi-warehouse inventory into parent groups with children."""
    groups = {}
    singles = []
    for entry in inv_data:
        grp = get_warehouse_group(entry['name'])
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)
    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        oh = sum(c['oh'] for c in children)
        rate = sum(c['rate'] for c in children)
        dep90 = sum(c['dep90'] for c in children)
        proj = sum(c['proj'] for c in children)
        doh = oh / rate if rate > 0 else 0
        status = compute_inventory_status_rep(doh)
        sku_agg = {}
        for c in children:
            for s in c.get('skus', []):
                if s['w'] not in sku_agg:
                    sku_agg[s['w']] = {'oh': 0, 'rate': 0}
                sku_agg[s['w']]['oh'] += s['oh']
                sku_agg[s['w']]['rate'] += s['rate']
        consolidated_skus = []
        for wine in WINE_ORDER:
            if wine in sku_agg:
                a = sku_agg[wine]
                s_doh = a['oh'] / a['rate'] if a['rate'] > 0 else 0
                consolidated_skus.append({
                    'w': wine, 'oh': round(a['oh'], 1),
                    'rate': round(a['rate'], 2), 'doh': round(s_doh),
                    'status': compute_inventory_status_rep(s_doh),
                })
        result.append({
            'name': grp_name, 'st': cfg['st'],
            'oh': round(oh, 1), 'rate': round(rate, 2),
            'doh': round(doh), 'dep90': round(dep90, 1),
            'status': status, 'proj': round(proj, 1),
            'skus': consolidated_skus,
            'isGroup': True,
            'children': sorted(children, key=lambda x: -x['oh']),
        })
    result.extend(singles)
    result.sort(key=lambda x: -x['oh'])
    return result

def consolidate_placements(placements):
    """Consolidate multi-warehouse placements into parent groups."""
    groups = {}
    singles = []
    for entry in placements:
        grp = get_warehouse_group(entry.get('fullName', ''))
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)
    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        result.append({
            'name': grp_name, 'st': cfg['st'],
            'newA': sum(c['newA'] for c in children),
            'reEngageA': sum(c['reEngageA'] for c in children),
            'net': sum(c['net'] for c in children),
            'isGroup': True,
            'children': children,
        })
    result.extend(singles)
    result.sort(key=lambda x: x['net'])
    return result


def build_new_wins(new_accounts):
    """Build new wins list (new placements)."""
    acct_agg = {}
    for a in new_accounts:
        key = a['account']
        if key not in acct_agg:
            acct_agg[key] = {
                'acct': a['account'], 'dist': a['dist'], 'st': a['state'],
                'ce': 0, 'skus': 0
            }
        acct_agg[key]['ce'] += a['sold_ce']
        acct_agg[key]['skus'] += a.get('new_accts', 1)

    all_accts = sorted(acct_agg.values(), key=lambda x: -x['ce'])
    return [{'acct': a['acct'], 'dist': a['dist'], 'st': a['st'],
             'ce': round(a['ce'], 1), 'skus': a['skus']} for a in all_accts]


def build_placement_summary_rep(new_dist, lost_dist):
    """Build placement summary (net new/lost by distributor, reframed as opportunities)."""
    all_dists = set(list(new_dist.keys()) + list(lost_dist.keys()))
    placements = []
    for dist in all_dists:
        nd = new_dist.get(dist, {})
        ld = lost_dist.get(dist, {})
        new_a = nd.get('new_accts', 0)
        re_engage_a = ld.get('lost_accts', 0)
        net = new_a - re_engage_a
        st = get_state_from_dist(dist)
        short = get_dist_short(dist)

        placements.append({
            'name': short, 'st': st,
            'newA': new_a,
            'reEngageA': re_engage_a,
            'net': net,
            'fullName': dist,
        })

    placements.sort(key=lambda x: x['net'])
    return placements


def build_accounts_top_rep(roll_accounts, roll_account_skus):
    """
    Build top accounts with reframed trends and growth potential.
    - Trend: "Momentum", "Growth Opportunity", "Consistent"
    - Growth Potential: if first half > second half, show opportunity gap
    """
    accounts = []
    for key, a in roll_accounts.items():
        if a['total_ce'] <= 0:
            continue

        nov = a['nov_ce']
        dec = a['dec_ce']
        jan = a['jan_ce']
        feb = a['feb_ce']
        first_half = nov + dec
        second_half = jan + feb

        trend = compute_trend_label(first_half, second_half)

        # Growth potential: gap between first and second half
        growth_potential = 0
        if first_half > second_half:
            growth_potential = round(first_half - second_half, 1)

        ch = a['channel'] if a['channel'] in ('ON', 'OFF') else 'OFF'

        accounts.append({
            'rank': 0,
            'acct': a['account'],
            'dist': a['dist'],
            'st': a['state'],
            'ch': ch,
            'nov': round(nov, 1),
            'dec': round(dec, 1),
            'jan': round(jan, 1),
            'feb': round(feb, 1),
            'total': round(a['total_ce'], 1),
            'trend': trend,
            'growthPotential': growth_potential,
        })

    accounts.sort(key=lambda x: -x['total'])
    for i, a in enumerate(accounts):
        a['rank'] = i + 1

    return accounts


def build_reorder_data_rep(roll_accounts, roll_account_sku_detail=None):
    """Build reorder forecast with SKU-level breakout (no revenue)."""
    today = datetime.now()
    reorder = []

    for key, a in roll_accounts.items():
        if a['total_ce'] <= 0 or not a['last_buy']:
            continue

        days_since = (today - a['last_buy']).days
        purchases = int(a['total_purchases'])

        if purchases > 1 and a['first_buy'] and a['last_buy'] and a['first_buy'] != a['last_buy']:
            span = (a['last_buy'] - a['first_buy']).days
            avg_cycle = span / (purchases - 1)
        else:
            avg_cycle = 30

        # Weighted priority score (0-100)
        # Overdue factor (40%): how far past cycle
        overdue_ratio = days_since / avg_cycle if avg_cycle > 0 else 1
        overdue_score = min(100, overdue_ratio * 50)
        # Volume factor (35%): higher CE = higher priority
        vol_score = min(100, a['total_ce'] * 5)
        # Recency factor (25%): more purchases = more reliable customer
        recency_score = min(100, purchases * 20)
        priority = round(overdue_score * 0.40 + vol_score * 0.35 + recency_score * 0.25)

        ch = a['channel'] if a['channel'] in ('ON', 'OFF') else 'OFF'

        # Build SKU-level breakout
        skus = []
        if roll_account_sku_detail:
            acct, dist = key
            for wine in WINE_ORDER:
                ask = (acct, dist, wine)
                asd = roll_account_sku_detail.get(ask)
                if not asd or asd['ce'] <= 0:
                    continue

                sku_purch = int(asd['total_purchases'])
                sku_days = 0
                sku_cycle = 0
                sku_last = ''

                if asd['last_buy']:
                    sku_days = (today - asd['last_buy']).days
                    sku_last = asd['last_buy'].strftime('%m/%d/%y')
                    if sku_purch > 1 and asd['first_buy'] and asd['first_buy'] != asd['last_buy']:
                        sku_span = (asd['last_buy'] - asd['first_buy']).days
                        sku_cycle = round(sku_span / (sku_purch - 1))
                    else:
                        sku_cycle = round(avg_cycle)
                else:
                    sku_cycle = round(avg_cycle)

                # Monthly pattern for this SKU
                months_active = sum(1 for m in [asd['nov_ce'], asd['dec_ce'], asd['jan_ce'], asd['feb_ce']] if m > 0)

                skus.append({
                    'w': wine,
                    'ce': round(asd['ce'], 1),
                    'purch': sku_purch,
                    'cycle': sku_cycle,
                    'last': sku_last,
                    'days': sku_days,
                    'months': months_active,
                })

        reorder.append({
            'rank': 0,
            'acct': a['account'],
            'dist': a['dist'],
            'st': a['state'],
            'ch': ch,
            'ce': round(a['total_ce'], 1),
            'purch': purchases,
            'cycle': round(avg_cycle),
            'last': a['last_buy'].strftime('%m/%d/%y'),
            'days': days_since,
            'priority': priority,
            'skus': skus,
        })

    reorder.sort(key=lambda x: -x['priority'])
    for i, r in enumerate(reorder):
        r['rank'] = i + 1

    return reorder


def build_dist_detail_rep(vel_dist, vel_dist_sku, four_w_dist):
    """
    Build distributor detail with reframed momentum (no revenue).
    """
    detail = []

    for dist in sorted(vel_dist.keys(), key=lambda d: -vel_dist[d].get('13w_ce', 0)):
        st = get_state_from_dist(dist)

        wines = []
        for wine in WINE_ORDER:
            sku = vel_dist_sku.get((dist, wine), {})
            ce = round(sku.get('13w_ce', 0), 1)
            if ce <= 0:
                continue

            prior = round(sku.get('prior_ce', 0), 1)
            momentum = fmt_momentum(prior, ce)

            # 4W approximation
            v = vel_dist.get(dist, {})
            fw = four_w_dist.get(dist, {})
            total_13w = v.get('13w_ce', 1)
            w4_total = fw.get('cur_ce', 0)
            w4 = round(w4_total * (ce / total_13w), 1) if total_13w > 0 else 0

            avg = round(ce / 13, 2)

            # Consistency
            sku_weekly = sku.get('weekly_ce', [0]*13)
            con = round(sum(1 for w in sku_weekly if w > 0) / 13, 2)

            wines.append({
                'w': wine,
                'ce': ce,
                'prior': prior,
                'momentum': momentum,
                'w4': w4,
                'avg': avg,
                'con': con,
            })

        if wines:
            entry = {
                'dist': f'{dist}, {st}' if st and f', {st}' not in dist else dist,
                'st': st,
                'wines': wines
            }
            detail.append(entry)

    return detail


def build_dist_health(vel_dist, vel_dist_sku, inv_dist, inv_dist_sku,
                      roll_accounts, roll_account_skus, dist_sku_detail,
                      new_dist, lost_dist, four_w_dist):
    """
    Build Distributor Health data: sell-in vs sell-through, inventory coverage,
    monthly purchasing patterns, and account penetration per distributor.
    """
    health = []

    for dist in sorted(vel_dist.keys(), key=lambda d: -vel_dist[d].get('13w_ce', 0)):
        st = get_state_from_dist(dist)
        dist_label = f'{dist}, {st}' if st and f', {st}' not in dist else dist

        v = vel_dist.get(dist, {})
        inv = inv_dist.get(dist, {})
        fw = four_w_dist.get(dist, {})

        # --- SKU-level sell-in vs sell-through + inventory ---
        sku_data = []
        for wine in WINE_ORDER:
            # Sell-through (depletions) from 13W velocity
            vel_sku = vel_dist_sku.get((dist, wine), {})
            sell_through = round(vel_sku.get('13w_ce', 0), 1)

            # Sell-in (purchases) from rolling period dist+sku aggregation
            ds = dist_sku_detail.get((dist, wine), {})
            sell_in = round(ds.get('ce', 0), 1)  # 4M total CE as sell-in proxy

            # Inventory per SKU -- try all states for this dist
            sku_oh = 0
            sku_rate = 0
            sku_doh = 0
            for (d, s, w), idata in inv_dist_sku.items():
                if d == dist and w == wine:
                    sku_oh += idata.get('on_hand', 0)
                    sku_rate += idata.get('daily_rate', 0)
                    sku_doh = idata.get('doh', 0)  # take last match (usually one state per dist)

            # Weeks of supply = on_hand / (weekly avg depletion)
            weekly_avg = sell_through / 13 if sell_through > 0 else 0
            weeks_supply = round(sku_oh / weekly_avg, 1) if weekly_avg > 0 else 0

            if sell_through <= 0 and sell_in <= 0 and sku_oh <= 0:
                continue

            # Sell-through ratio
            ratio = round((sell_through / sell_in) * 100, 0) if sell_in > 0 else 0

            sku_data.append({
                'w': wine,
                'sellIn': sell_in,
                'sellThru': sell_through,
                'ratio': ratio,
                'oh': round(sku_oh, 1),
                'wkSupply': weeks_supply,
            })

        # --- Monthly purchasing pattern (aggregate all accounts for this dist) ---
        nov = dec = jan = feb = 0
        total_accounts = 0
        active_accounts = 0
        sku_breadth_sum = 0
        engagement_counts = {'Established': 0, 'Building': 0, 'Emerging': 0}

        for (acct, d), a in roll_accounts.items():
            if d != dist:
                continue
            total_accounts += 1
            nov += a.get('nov_ce', 0)
            dec += a.get('dec_ce', 0)
            jan += a.get('jan_ce', 0)
            feb += a.get('feb_ce', 0)

            acct_ce = a.get('total_ce', 0)
            if acct_ce > 0:
                active_accounts += 1

            sku_count = a.get('sku_count', 0)
            sku_breadth_sum += sku_count

            # Engagement level based on SKU breadth and volume
            months_active = sum(1 for m in [a.get('nov_ce',0), a.get('dec_ce',0), a.get('jan_ce',0), a.get('feb_ce',0)] if m > 0)
            if sku_count >= 3 and months_active >= 3:
                engagement_counts['Established'] += 1
            elif sku_count >= 2 or months_active >= 2:
                engagement_counts['Building'] += 1
            elif acct_ce > 0:
                engagement_counts['Emerging'] += 1

        avg_sku_breadth = round(sku_breadth_sum / active_accounts, 1) if active_accounts > 0 else 0

        # Net new placements
        nd = new_dist.get(dist, {})
        ld = lost_dist.get(dist, {})
        new_accts = nd.get('new_accts', 0)
        lost_accts = ld.get('lost_accts', 0)

        # Total sell-in / sell-through at dist level
        total_sell_in = round(sum(s['sellIn'] for s in sku_data), 1)
        total_sell_thru = round(sum(s['sellThru'] for s in sku_data), 1)
        total_ratio = round((total_sell_thru / total_sell_in) * 100, 0) if total_sell_in > 0 else 0

        entry = {
            'dist': dist_label,
            'st': st,
            'totalSellIn': total_sell_in,
            'totalSellThru': total_sell_thru,
            'totalRatio': total_ratio,
            'oh': round(inv.get('on_hand', 0), 1),
            'doh': round(inv.get('doh', 0)),
            'skus': sku_data,
            'nov': round(nov, 1),
            'dec': round(dec, 1),
            'jan': round(jan, 1),
            'feb': round(feb, 1),
            'totalAccounts': total_accounts,
            'activeAccounts': active_accounts,
            'avgSkuBreadth': avg_sku_breadth,
            'established': engagement_counts['Established'],
            'building': engagement_counts['Building'],
            'emerging': engagement_counts['Emerging'],
            'newAccts': new_accts,
            'lostAccts': lost_accts,
        }
        health.append(entry)

    return health


def consolidate_dist_health(health):
    """Consolidate multi-warehouse distributor health entries into parent groups."""
    groups = {}
    singles = []
    for entry in health:
        grp = get_warehouse_group(entry['dist'])
        if grp:
            groups.setdefault(grp, []).append(entry)
        else:
            singles.append(entry)

    result = []
    for grp_name, children in groups.items():
        cfg = MULTI_WAREHOUSE_GROUPS[grp_name]
        st = cfg['st']

        # Merge SKU data across warehouses
        sku_agg = {}
        for c in children:
            for s in c.get('skus', []):
                w = s['w']
                if w not in sku_agg:
                    sku_agg[w] = {'sellIn': 0, 'sellThru': 0, 'oh': 0, 'wkSupply': 0}
                sku_agg[w]['sellIn'] += s['sellIn']
                sku_agg[w]['sellThru'] += s['sellThru']
                sku_agg[w]['oh'] += s['oh']

        sku_data = []
        for w in WINE_ORDER:
            if w not in sku_agg:
                continue
            a = sku_agg[w]
            ratio = round((a['sellThru'] / a['sellIn']) * 100, 0) if a['sellIn'] > 0 else 0
            weekly_avg = a['sellThru'] / 13 if a['sellThru'] > 0 else 0
            wk_supply = round(a['oh'] / weekly_avg, 1) if weekly_avg > 0 else 0
            sku_data.append({
                'w': w,
                'sellIn': round(a['sellIn'], 1),
                'sellThru': round(a['sellThru'], 1),
                'ratio': ratio,
                'oh': round(a['oh'], 1),
                'wkSupply': wk_supply,
            })

        total_sell_in = round(sum(s['sellIn'] for s in sku_data), 1)
        total_sell_thru = round(sum(s['sellThru'] for s in sku_data), 1)
        total_ratio = round((total_sell_thru / total_sell_in) * 100, 0) if total_sell_in > 0 else 0

        result.append({
            'dist': grp_name,
            'st': st,
            'totalSellIn': total_sell_in,
            'totalSellThru': total_sell_thru,
            'totalRatio': total_ratio,
            'oh': round(sum(c['oh'] for c in children), 1),
            'doh': round(sum(c['doh'] * c['oh'] for c in children) / max(sum(c['oh'] for c in children), 0.01)),
            'skus': sku_data,
            'nov': round(sum(c['nov'] for c in children), 1),
            'dec': round(sum(c['dec'] for c in children), 1),
            'jan': round(sum(c['jan'] for c in children), 1),
            'feb': round(sum(c['feb'] for c in children), 1),
            'totalAccounts': sum(c['totalAccounts'] for c in children),
            'activeAccounts': sum(c['activeAccounts'] for c in children),
            'avgSkuBreadth': round(
                sum(c['avgSkuBreadth'] * c['activeAccounts'] for c in children) /
                max(sum(c['activeAccounts'] for c in children), 1), 1),
            'established': sum(c['established'] for c in children),
            'building': sum(c['building'] for c in children),
            'emerging': sum(c['emerging'] for c in children),
            'newAccts': sum(c['newAccts'] for c in children),
            'lostAccts': sum(c['lostAccts'] for c in children),
        })

    result.extend(singles)
    result.sort(key=lambda x: -x['totalSellThru'])
    return result


# ============================================================
# JAVASCRIPT DATA GENERATION
# ============================================================

def generate_js_data_rep(scorecard, inventory, reeng, new_wins, placement_summary,
                         accounts_top, reorder, dist_detail, dist_health, week_headers=None,
                         qb_dist_orders=None, warehouse_inv=None, classic_tracker=None,
                         sample_accounts=None, sample_placements=None):
    """Generate JavaScript data for rep dashboard (no revenue, opportunity-focused)."""

    lines = []

    # State names
    states_found = set()
    for s in scorecard:
        if s['st']:
            states_found.add(s['st'])
    for s in inventory:
        if s['st']:
            states_found.add(s['st'])

    state_entries = ','.join(f"{st}:'{STATE_NAMES.get(st, st)}'" for st in sorted(states_found) if st in STATE_NAMES)
    lines.append(f'const stateNames = {{\n  {state_entries}\n}};')
    lines.append('')

    lines.append(f'const regionMap = {json.dumps(REGION_MAP)};')
    lines.append('')

    # ===== DEPLETION SCORECARD (no revenue) =====
    lines.append('// ==================== DEPLETION SCORECARD ====================')
    lines.append('const distScorecard = [')
    for s in scorecard:
        weekly_str = ','.join(str(w) for w in s['weekly'])
        children_str = ''
        if s.get('isGroup'):
            children_list = []
            for child in s.get('children', []):
                child_weekly = ','.join(str(w) for w in child['weekly'])
                children_list.append(f"{{name:'{js_str(child['name'])}',st:'{child['st']}',ce:{child['ce']},prior:{child['prior']},momentum:'{child['momentum']}',w4:{child['w4']},prior_w4:{child.get('prior_w4',0)},w4trend:'{child['w4trend']}',oh:{child['oh']},doh:{child['doh']},net:{child['net']},con:{child['con']},conLabel:'{child['conLabel']}',invAction:'{child.get('invAction','')}',velTrend:'{child.get('velTrend','')}',sellThru:{child.get('sellThru',0)},weekly:[{child_weekly}]}}")
            children_str = f",children:[{','.join(children_list)}]"
        lines.append(f"  {{name:'{js_str(s['name'])}',st:'{s['st']}',ce:{s['ce']},prior:{s['prior']},momentum:'{s['momentum']}',w4:{s['w4']},prior_w4:{s.get('prior_w4',0)},w4trend:'{s['w4trend']}',oh:{s['oh']},doh:{s['doh']},net:{s['net']},con:{s['con']},conLabel:'{s['conLabel']}',invAction:'{s.get('invAction','')}',velTrend:'{s.get('velTrend','')}',sellThru:{s.get('sellThru',0)},")
        lines.append(f"   weekly:[{weekly_str}],isGroup:{str(s.get('isGroup',False)).lower()}{children_str}}},")
    lines.append('];')
    lines.append('')

    # ===== INVENTORY DATA (reframed status, with SKU breakout) =====
    lines.append('const inventoryData = [')
    for inv in inventory:
        sku_parts = []
        for s in inv.get('skus', []):
            sku_parts.append(f"{{w:'{s['w']}',oh:{s['oh']},rate:{s['rate']},doh:{s['doh']},status:'{s['status']}'}}")
        skus_str = ','.join(sku_parts)
        children_str = ''
        if inv.get('isGroup'):
            children_list = []
            for child in inv.get('children', []):
                child_skus = []
                for s in child.get('skus', []):
                    child_skus.append(f"{{w:'{s['w']}',oh:{s['oh']},rate:{s['rate']},doh:{s['doh']},status:'{s['status']}'}}")
                child_skus_str = ','.join(child_skus)
                children_list.append(f"{{name:'{js_str(child['name'])}',st:'{child['st']}',oh:{child['oh']},rate:{child['rate']},doh:{child['doh']},dep90:{child['dep90']},status:'{child['status']}',proj:{child['proj']},skus:[{child_skus_str}]}}")
            children_str = f",children:[{','.join(children_list)}]"
        lines.append(f"  {{name:'{js_str(inv['name'])}',st:'{inv['st']}',oh:{inv['oh']},rate:{inv['rate']},doh:{inv['doh']},dep90:{inv['dep90']},status:'{inv['status']}',proj:{inv['proj']},skus:[{skus_str}],isGroup:{str(inv.get('isGroup',False)).lower()}{children_str}}},")
    lines.append('];')
    lines.append('')

    # ===== RE-ENGAGEMENT OPPORTUNITIES =====
    lines.append('// ==================== RE-ENGAGEMENT OPPORTUNITIES ====================')
    lines.append('const reEngagementData = [')
    for r in reeng:
        lines.append(f"  {{name:'{js_str(r['account'])}',st:'{r['state']}',priorAccts:{r['priorAccts']},priorCE:{r['priorCE']},priority:{r['priority']}}},")
    lines.append('];')
    lines.append('')

    # ===== NEW WINS =====
    lines.append('const newWins = [')
    for w in new_wins:
        lines.append(f"  {{acct:'{js_str(w['acct'])}',dist:'{js_str(w['dist'])}',st:'{w['st']}',ce:{w['ce']},skus:{w['skus']}}},")
    lines.append('];')
    lines.append('')

    # ===== PLACEMENT SUMMARY (reframed) =====
    lines.append('const placementSummary = [')
    for p in placement_summary:
        children_str = ''
        if p.get('isGroup'):
            children_list = []
            for child in p.get('children', []):
                children_list.append(f"{{name:'{js_str(child['name'])}',st:'{child['st']}',newA:{child['newA']},reEngageA:{child['reEngageA']},net:{child['net']}}}")
            children_str = f",children:[{','.join(children_list)}]"
        lines.append(f"  {{name:'{js_str(p['name'])}',st:'{p['st']}',newA:{p['newA']},reEngageA:{p['reEngageA']},net:{p['net']},isGroup:{str(p.get('isGroup',False)).lower()}{children_str}}},")
    lines.append('];')
    lines.append('')

    # ===== SAMPLE / MARKETING DATA =====
    lines.append('// ==================== SAMPLE / MARKETING DATA ====================')
    sample_by_dist = defaultdict(lambda: {'ce': 0, 'rev': 0, 'count': 0})
    if sample_accounts:
        for (acct, dist), sdata in sample_accounts.items():
            sample_by_dist[dist]['ce'] += sdata.get('total_ce', 0)
            sample_by_dist[dist]['rev'] += sdata.get('total_rev', 0)
            sample_by_dist[dist]['count'] += 1
    if sample_placements:
        for sp in sample_placements:
            sample_by_dist[sp['dist']]['placement_ce'] = sample_by_dist[sp['dist']].get('placement_ce', 0) + sp.get('sold_ce', 0)
            sample_by_dist[sp['dist']]['placement_count'] = sample_by_dist[sp['dist']].get('placement_count', 0) + sp.get('new_accts', 0)

    sample_list = []
    for dist, sd in sorted(sample_by_dist.items(), key=lambda x: -x[1]['ce']):
        st = ''
        for entry in scorecard:
            if entry['name'] == dist:
                st = entry.get('st', '')
                break
        sample_list.append(f"{{dist:'{js_str(dist)}',st:'{st}',ce:{round(sd['ce'],1)},rev:{round(sd.get('rev',0),2)},accts:{sd['count']},placeCE:{round(sd.get('placement_ce',0),1)},placeCount:{sd.get('placement_count',0)}}}")
    lines.append(f"const sampleSummary = [{','.join(sample_list)}];")
    lines.append('')

    # ===== TOP ACCOUNTS (with growth potential) =====
    lines.append('// ==================== TOP ACCOUNTS ====================')
    lines.append('const accountsTop = [')
    for a in accounts_top:
        lines.append(f"  {{rank:{a['rank']},acct:'{js_str(a['acct'])}',dist:'{js_str(a['dist'])}',st:'{a['st']}',ch:'{a['ch']}',nov:{a['nov']},dec:{a['dec']},jan:{a['jan']},feb:{a['feb']},total:{a['total']},trend:'{a['trend']}',growthPotential:{a['growthPotential']}}},")
    lines.append('];')
    lines.append('')

    # ===== ACCOUNT CONCENTRATION =====
    lines.append('// ==================== ACCOUNT CONCENTRATION ====================')
    if accounts_top:
        all_totals = sorted([a['total'] for a in accounts_top], reverse=True)
        grand_total = sum(all_totals)
        if grand_total > 0:
            top5_pct = round(sum(all_totals[:5]) / grand_total * 100, 1) if len(all_totals) >= 5 else 0
            top10_pct = round(sum(all_totals[:10]) / grand_total * 100, 1) if len(all_totals) >= 10 else 0
            top20_pct = round(sum(all_totals[:20]) / grand_total * 100, 1) if len(all_totals) >= 20 else 0
            median_ce = round(all_totals[len(all_totals)//2], 1)
            under1ce = sum(1 for t in all_totals if t < 1.0)
        else:
            top5_pct = top10_pct = top20_pct = median_ce = 0
            under1ce = 0
        lines.append(f'const acctConcentration = {{total:{len(all_totals)},top5:{top5_pct},top10:{top10_pct},top20:{top20_pct},median:{median_ce},under1:{under1ce}}};')
    else:
        lines.append('const acctConcentration = {total:0,top5:0,top10:0,top20:0,median:0,under1:0};')
    lines.append('')

    # ===== REORDER FORECAST (with SKU breakout) =====
    lines.append('const reorderData = [')
    for r in reorder:
        sku_parts = []
        for s in r.get('skus', []):
            sku_parts.append(f"{{w:'{s['w']}',ce:{s['ce']},purch:{s['purch']},cycle:{s['cycle']},last:'{s['last']}',days:{s['days']},months:{s['months']}}}")
        skus_str = ','.join(sku_parts)
        lines.append(f"  {{rank:{r['rank']},acct:'{js_str(r['acct'])}',dist:'{js_str(r['dist'])}',st:'{r['st']}',ch:'{r['ch']}',ce:{r['ce']},purch:{r['purch']},cycle:{r['cycle']},last:'{r['last']}',days:{r['days']},priority:{r['priority']},skus:[{skus_str}]}},")
    lines.append('];')
    lines.append('')

    # ===== DISTRIBUTOR DETAIL (reframed momentum, no revenue) =====
    lines.append('// ==================== DISTRIBUTOR DETAIL ====================')
    lines.append('const distDetail = [')
    for d in dist_detail:
        lines.append(f"  {{dist:'{js_str(d['dist'])}',st:'{d['st']}',wines:[")
        for w in d['wines']:
            lines.append(f"    {{w:'{w['w']}',ce:{w['ce']},prior:{w['prior']},momentum:'{w['momentum']}',w4:{w['w4']},avg:{w['avg']},con:{w['con']}}},")
        lines.append('  ]},')
    lines.append('];')
    lines.append('')

    # ===== DISTRIBUTOR HEALTH =====
    lines.append('// ==================== DISTRIBUTOR HEALTH ====================')
    lines.append('const distHealth = [')
    for h in dist_health:
        sku_lines = []
        for s in h['skus']:
            sku_lines.append(f"{{w:'{s['w']}',sellIn:{s['sellIn']},sellThru:{s['sellThru']},ratio:{s['ratio']},oh:{s['oh']},wkSupply:{s['wkSupply']}}}")
        skus_str = ','.join(sku_lines)
        lines.append(f"  {{dist:'{js_str(h['dist'])}',st:'{h['st']}',totalSellIn:{h['totalSellIn']},totalSellThru:{h['totalSellThru']},totalRatio:{h['totalRatio']},oh:{h['oh']},doh:{h['doh']},")
        lines.append(f"   skus:[{skus_str}],")
        lines.append(f"   nov:{h['nov']},dec:{h['dec']},jan:{h['jan']},feb:{h['feb']},")
        lines.append(f"   totalAccounts:{h['totalAccounts']},activeAccounts:{h['activeAccounts']},avgSkuBreadth:{h['avgSkuBreadth']},")
        lines.append(f"   established:{h['established']},building:{h['building']},emerging:{h['emerging']},newAccts:{h['newAccts']},lostAccts:{h['lostAccts']}}},")
    lines.append('];')
    lines.append('')

    # ===== BUILD METADATA =====
    # ===== QB DISTRIBUTOR ORDERS =====
    lines.append('// ==================== QB DISTRIBUTOR ORDERS ====================')
    lines.append('const qbDistOrders = {')
    if qb_dist_orders:
        for dist_name, info in sorted(qb_dist_orders.items()):
            lines.append(f"  '{js_str(dist_name)}':{{date:'{info['last_date']}',num:'{info['last_num']}',total:{info['last_total']},items:'{js_str(info['last_items'])}',orders:{info['order_count']},label:'{info.get('label_status','Unknown')}',classicCases:{info.get('classic_cases',0)},contemporaryCases:{info.get('contemporary_cases',0)}}},")
    lines.append('};')
    lines.append('')

    # ===== WAREHOUSE INVENTORY =====
    lines.append('// ==================== WAREHOUSE INVENTORY ====================')
    if warehouse_inv:
        lines.append('const warehouseInventory = {')
        for section in ['classic', 'contemporary']:
            lines.append(f"  {section}: {{")
            for loc, vals in warehouse_inv.get(section, {}).items():
                sw = vals.get('Still White', 0)
                sr = vals.get('Still Rose', 0)
                rd = vals.get('Red', 0)
                spw = vals.get('Sparkling White', 0)
                spr = vals.get('Sparkling Rose', 0)
                t = vals.get('total', 0)
                lines.append(f"    '{js_str(loc)}':{{sw:{sw},sr:{sr},rd:{rd},spw:{spw},spr:{spr},total:{t}}},")
            lines.append('  },')
        gt = warehouse_inv.get('grand_total', {})
        lines.append(f"  grandTotal:{{sw:{gt.get('Still White',0)},sr:{gt.get('Still Rose',0)},rd:{gt.get('Red',0)},spw:{gt.get('Sparkling White',0)},spr:{gt.get('Sparkling Rose',0)},total:{gt.get('total',0)}}},")
        b4 = warehouse_inv.get('batch4_estimates', {})
        lines.append(f"  batch4Est:{{sw:{b4.get('Still White',0)},sr:{b4.get('Still Rose',0)},rd:{b4.get('Red',0)},spw:{b4.get('Sparkling White',0)},spr:{b4.get('Sparkling Rose',0)},total:{b4.get('total',0)}}},")
        lines.append('};')
    else:
        lines.append('const warehouseInventory = null;')
    lines.append('')

    # Classic Tracker
    ct = classic_tracker
    if ct:
        lines.append('const classicTracker = {')
        lines.append(f"  trackable: {ct['classic_trackable']},")
        lines.append(f"  totalWithBad: {ct['classic_total']},")
        lines.append(f"  badSpr: {ct['bad_spr']},")
        lines.append(f"  burnRate: {ct['burn_rate_monthly']},")
        lines.append(f"  monthsRemaining: {ct['months_remaining']},")
        lines.append(f"  requiredMonthly: {ct['required_monthly']},")
        lines.append(f"  dataDate: '{ct['data_date']}',")
        # By-wine breakdown
        bw = ct['classic_by_wine']
        lines.append(f"  byWine: {{sw:{bw.get('Still White',0)},sr:{bw.get('Still Rose',0)},rd:{bw.get('Red',0)},spw:{bw.get('Sparkling White',0)},spr:{bw.get('Sparkling Rose',0)}}},")
        # Location details
        lines.append('  locations: {')
        for loc, vals in ct['classic_locations'].items():
            sw = round(vals.get('Still White', 0), 1)
            sr = round(vals.get('Still Rose', 0), 1)
            rd = round(vals.get('Red', 0), 1)
            spw = round(vals.get('Sparkling White', 0), 1)
            spr = round(vals.get('Sparkling Rose', 0), 1)
            t = round(vals.get('total', 0), 1)
            lines.append(f"    '{js_str(loc)}':{{sw:{sw},sr:{sr},rd:{rd},spw:{spw},spr:{spr},total:{t}}},")
        lines.append('  },')
        # History for burn-down chart
        lines.append('  history: [')
        for h in ct['history']:
            lines.append(f"    {{date:'{h['date']}',total:{round(h['total'],1)}}},")
        lines.append('  ],')
        # New label totals for context
        lines.append(f"  newLabelTotal: {ct['new_label_total']},")
        lines.append('};')
    else:
        lines.append('const classicTracker = null;')
    lines.append('')

    lines.append('// ==================== BUILD METADATA ====================')
    build_date = datetime.now().strftime('%B %d, %Y')
    last_week = week_headers[-1] if week_headers else ''
    lines.append(f"const buildDate = '{build_date}';")
    lines.append(f"const dataThrough = '{last_week}';")

    return '\n'.join(lines)


# ============================================================
# HTML INJECTION
# ============================================================

def inject_data_into_html(template_path, js_data, output_path):
    """Replace data section in HTML template with fresh data."""
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    start_marker = '// __DATA_START__'
    end_marker = '// __DATA_END__'

    start_idx = html.find(start_marker)
    end_idx = html.find(end_marker)

    if start_idx == -1 or end_idx == -1:
        print("ERROR: Data markers not found in HTML template!")
        print("Make sure the template has '// __DATA_START__' and '// __DATA_END__' markers.")
        sys.exit(1)

    new_html = html[:start_idx] + start_marker + '\n' + js_data + '\n' + end_marker + html[end_idx + len(end_marker):]

    # Update date in header
    today_str = datetime.now().strftime('%b %d, %Y')
    new_html = re.sub(
        r'Data as of <strong>[^<]+</strong>',
        f'Data as of <strong>{today_str}</strong>',
        new_html
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(new_html)

    print(f"Rep dashboard written to: {output_path}")
    return output_path


# ============================================================
# MAIN
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Rebuild Missing Thorn REP Dashboard from Excel exports')
    parser.add_argument('--data', '--base-dir', dest='data', default=None,
                        help='Path to directory containing Excel source files')
    parser.add_argument('--template', default=None, help='Path to HTML template')
    parser.add_argument('--output', default=None, help='Output path for generated HTML')
    args = parser.parse_args()

    # Determine paths relative to repo root (one level up from scripts/)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(script_dir)

    # Data directory: --data flag > data/ in repo root > same dir as script (legacy)
    base_dir = args.data
    if not base_dir:
        repo_data = os.path.join(repo_root, 'data')
        if os.path.isdir(repo_data):
            base_dir = repo_data
        else:
            base_dir = script_dir
    if not os.path.isdir(base_dir):
        print(f"ERROR: Data directory not found: {base_dir}")
        print("Provide --data flag pointing to your Excel files directory.")
        sys.exit(1)

    # Template: --template flag > templates/rep_dashboard_template.html in repo root
    template = args.template
    if not template:
        repo_template = os.path.join(repo_root, 'templates', 'rep_dashboard_template.html')
        if os.path.exists(repo_template):
            template = repo_template
        else:
            # Legacy: look for most recent *_MT_BI_RepDashboard.html in script dir
            candidates = [f for f in os.listdir(script_dir) if f.endswith('_MT_BI_RepDashboard.html')]
            if candidates:
                template = os.path.join(script_dir, sorted(candidates)[-1])

    if not template or not os.path.exists(template):
        print(f"ERROR: No HTML template found. Provide --template path.")
        sys.exit(1)

    print(f"Data directory: {base_dir}")
    print(f"Template: {template}")

    # ---- Load all Excel data ----
    print("Loading 13W velocity data...")
    vel_grand, vel_dist, vel_dist_sku, week_headers = load_13w_velocity(
        os.path.join(base_dir, 'MT_13W_Distributor_SKU_Velocity.xlsx'))

    print("Loading 4W comparison data...")
    four_w_grand, four_w_dist = load_4w_comparison(
        os.path.join(base_dir, 'MT_4W_vs_Prior4W_SKU.xlsx'))

    print("Loading inventory data...")
    inv_dist, inv_dist_sku = load_inventory(
        os.path.join(base_dir, 'MT_Inventory_DaysOnHand.xlsx'))

    print("Loading placement data...")
    new_dist, new_accounts_raw, lost_dist, lost_accounts_raw, sample_placements = load_placements(base_dir)

    print("Loading 4M rolling period data...")
    roll_accounts, roll_account_skus, dist_sku_detail, roll_account_sku_detail, sample_accounts = load_rolling_period(
        os.path.join(base_dir, 'MT_4M_RollingPeriod.xlsx'))

    print(f"  Samples excluded: {len(sample_placements)} placement entries, {len(sample_accounts)} account entries")

    print("Loading QuickBooks distributor orders...")
    qb_file = os.path.join(base_dir, 'Sales by Customer Detail - RD.xlsx')
    if not os.path.exists(qb_file):
        alt = os.path.join(base_dir, 'YOLO_SalesByCustomerDetail.xlsx')
        if os.path.exists(alt):
            qb_file = alt
    qb_dist_orders = {}
    if os.path.exists(qb_file):
        raw_qb_orders = load_qb_distributor_orders(qb_file)
        # Get all depletion distributor names
        dep_dist_names = list(set(a['dist'] for a in roll_accounts.values()))
        # Match QB names to depletion names
        for qb_name, order_info in raw_qb_orders.items():
            matched = match_qb_dist_name(qb_name, dep_dist_names)
            for dep_name in matched:
                qb_dist_orders[dep_name] = order_info
        print(f"  Matched {len(qb_dist_orders)} distributors to QB orders")
    else:
        print("  QB file not found, skipping order data")

    print("Loading warehouse inventory...")
    inv_tracker_file_wh = os.path.join(base_dir, 'Inventory Tracker Live.xlsx')
    warehouse_inv = None
    if os.path.exists(inv_tracker_file_wh):
        warehouse_inv = load_warehouse_inventory(inv_tracker_file_wh)
        classic_total = warehouse_inv.get('classic', {}).get('Warehouse Total', {}).get('total', 0)
        contemp_total = warehouse_inv.get('contemporary', {}).get('Warehouse Total', {}).get('total', 0)
        print(f"  Classic stock: {classic_total:,.0f} cases | Contemporary stock: {contemp_total:,.0f} cases")
        print(f"  Data from tab: {warehouse_inv.get('data_date', 'unknown')}")
    else:
        print("  Inventory Tracker Live.xlsx not found, skipping warehouse data")

    print("Loading Classic Sellout Tracker...")
    classic_tracker = None
    inv_tracker_file = os.path.join(base_dir, 'Inventory Tracker Live.xlsx')
    if os.path.exists(inv_tracker_file):
        classic_tracker = load_inventory_tracker_live(inv_tracker_file)
        print(f"  Classic trackable: {classic_tracker['classic_trackable']:,.0f} cases | Burn rate: {classic_tracker['burn_rate_monthly']:.1f}/mo")
    else:
        print("  Inventory Tracker Live.xlsx not found, skipping classic tracker")

    # ---- Transform data (REP-FOCUSED) ----
    print("Building rep dashboard data...")
    scorecard = build_dist_scorecard_rep(vel_dist, four_w_dist, inv_dist, new_dist, lost_dist)
    scorecard = consolidate_scorecard(scorecard)
    inventory = build_inventory_data_rep(inv_dist, inv_dist_sku)
    inventory = consolidate_inventory(inventory)
    reeng = build_reengagement_data(lost_accounts_raw)
    new_wins = build_new_wins(new_accounts_raw)
    placement_summary = build_placement_summary_rep(new_dist, lost_dist)
    placement_summary = consolidate_placements(placement_summary)
    accounts_top = build_accounts_top_rep(roll_accounts, roll_account_skus)
    reorder = build_reorder_data_rep(roll_accounts, roll_account_sku_detail)
    dist_detail = build_dist_detail_rep(vel_dist, vel_dist_sku, four_w_dist)
    dist_health = build_dist_health(vel_dist, vel_dist_sku, inv_dist, inv_dist_sku,
                                     roll_accounts, roll_account_skus, dist_sku_detail,
                                     new_dist, lost_dist, four_w_dist)
    dist_health = consolidate_dist_health(dist_health)

    # ---- Generate JS data ----
    print("Generating JavaScript data...")
    js_data = generate_js_data_rep(
        scorecard, inventory, reeng, new_wins, placement_summary,
        accounts_top, reorder, dist_detail, dist_health, week_headers,
        qb_dist_orders, warehouse_inv, classic_tracker, sample_accounts, sample_placements)

    # ---- Inject into HTML ----
    output = args.output
    if not output:
        output = os.path.join(repo_root, 'dist', 'index_rep.html')

    os.makedirs(os.path.dirname(output) or '.', exist_ok=True)
    result = inject_data_into_html(template, js_data, output)

    print(f"\nRep dashboard rebuild complete!")
    print(f"  Distributors: {len(scorecard)}")
    print(f"  Inventory entries: {len(inventory)}")
    print(f"  Re-engagement opportunities: {len(reeng)}")
    print(f"  New wins: {len(new_wins)}")
    print(f"  Top accounts: {len(accounts_top)}")
    print(f"  Reorder forecast: {len(reorder)}")


if __name__ == '__main__':
    main()
